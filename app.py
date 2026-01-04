from db import db
import os
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
from template_updater import AssessmentTemplateUpdater

import io
import csv
import random
from functools import wraps
from werkzeug.utils import secure_filename
from datetime import datetime

from flask import Flask, render_template, redirect, url_for, flash, request, send_file, abort, jsonify
from flask_login import LoginManager, login_user, login_required, logout_user, current_user
from flask_bcrypt import Bcrypt
from flask_wtf import FlaskForm
from flask_wtf.file import FileField, FileAllowed
from wtforms import StringField, PasswordField, FloatField, SelectField, SelectMultipleField, DateField, TextAreaField, BooleanField
from wtforms.validators import InputRequired, Length, Optional, NumberRange

from config import config
from models import User, Student, Assessment, Setting, ActivityLog, Question, QuestionAttempt, Quiz, QuizAttempt, init_db
from excel_utils import ExcelTemplateHandler, ExcelBulkImporter, StudentBulkImporter, create_default_template, create_student_import_template, create_question_import_template

# -------------------------
# Application Factory
# -------------------------
app = Flask(__name__, static_folder='public')

# Load configuration
env = os.environ.get('FLASK_ENV', 'development')
app.config.from_object(config[env])

# File upload configuration
app.config['UPLOAD_FOLDER'] = os.path.join(os.path.dirname(__file__), 'uploads')
app.config['TEMPLATE_FOLDER'] = os.path.join(os.path.dirname(__file__), 'templates_excel')
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size

# Create necessary folders
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['TEMPLATE_FOLDER'], exist_ok=True)

# -------------------------
# Extensions
# -------------------------
bcrypt = Bcrypt(app)
login_manager = LoginManager(app)
login_manager.login_view = "login"

# Initialize database
init_db(app, bcrypt)

# -------------------------
# Activity Logging
# -------------------------
def log_activity(user, action, details=None):
    """Log user activity for auditing purposes"""
    if not user or not user.is_authenticated:
        return
    try:
        ip_address = request.remote_addr if request else None
        log_entry = ActivityLog(
            user_id=user.id,
            action=action,
            details=details,
            ip_address=ip_address
        )
        db.session.add(log_entry)
        db.session.commit()
    except Exception as e:
        # Log to console if database logging fails
        print(f"Failed to log activity: {e}")

# -------------------------
# Forms - FIXED: Remove duplicate definitions
# -------------------------
# Forms - FIXED: Remove duplicate definitions
# -------------------------

class StudentLoginForm(FlaskForm):
    first_name = StringField("First Name", validators=[InputRequired(), Length(min=1, max=120)])
    student_number = StringField("Student Number", validators=[InputRequired(), Length(min=1, max=50)])

class LoginForm(FlaskForm):
    username = StringField("Username", validators=[InputRequired(), Length(min=3, max=80)])
    password = PasswordField("Password", validators=[InputRequired(), Length(min=4)])

class UserForm(FlaskForm):
    username = StringField("Username", validators=[InputRequired(), Length(min=3)])
    password = PasswordField("Password", validators=[InputRequired(), Length(min=6)])
    role = SelectField("Role", choices=app.config['USER_ROLES'])
    subject = SelectField("Subject (for teachers)", choices=[("", "-- Not Applicable --")] + app.config['LEARNING_AREAS'], validators=[Optional()])
    class_name = SelectField("Class (for teachers)", choices=[("", "-- Not Applicable --")] + app.config['CLASS_LEVELS'], validators=[Optional()])

class EditUserForm(FlaskForm):
    role = SelectField("Role", choices=app.config['USER_ROLES'])
    subject = SelectField("Subject (for teachers)", choices=[("", "-- Not Applicable --")] + app.config['LEARNING_AREAS'], validators=[Optional()])
    class_name = SelectField("Class (for teachers)", choices=[("", "-- Not Applicable --")] + app.config['CLASS_LEVELS'], validators=[Optional()])

class PasswordResetForm(FlaskForm):
    password = PasswordField("New Password", validators=[InputRequired(), Length(min=6)])

# StudentForm - ONE DEFINITION ONLY
class StudentForm(FlaskForm):
    student_number = StringField("Student Number", validators=[InputRequired(), Length(min=1, max=50)])
    first_name = StringField("First name", validators=[InputRequired()])
    last_name = StringField("Last name", validators=[InputRequired()])
    middle_name = StringField("Middle name", validators=[Optional()])
    class_name = SelectField("Class", choices=[("", "-- Select Class --")] + app.config['CLASS_LEVELS'], validators=[Optional()])
    study_area = SelectField("Study/Learning Area", choices=[("", "-- Select Study Area --")] + app.config['STUDY_AREAS'], validators=[Optional()])
    # Removed duplicate learning_area field since it's not in your models

class AssessmentForm(FlaskForm):
    student_number = StringField("Student Number", validators=[Optional()])
    student_name = SelectField("Student Name", choices=[], validators=[InputRequired()])
    reference_number = StringField("Reference Number", validators=[Optional()])
    category = SelectField("Category", choices=app.config['ASSESSMENT_CATEGORIES'], validators=[InputRequired()])
    subject = SelectField("Subject", choices=[("", "-- Select Subject --")] + app.config['LEARNING_AREAS'], validators=[InputRequired()])
    class_name = SelectField("Class", choices=[("", "-- Select Class --")] + app.config['CLASS_LEVELS'], validators=[Optional()])
    score = FloatField("Score", validators=[InputRequired(), NumberRange(min=0)])
    max_score = SelectField("Max Score", choices=[(50, '50'), (100, '100')], validators=[InputRequired()], default=100)
    term = SelectField("Term", choices=app.config['TERMS'], validators=[InputRequired()])
    academic_year = StringField("Academic Year", validators=[Optional()])
    session = StringField("Session", validators=[Optional()])
    assessor = StringField("Assessor", validators=[Optional()])
    comments = TextAreaField("Comments", validators=[Optional()])

class TeacherAssignmentForm(FlaskForm):
    subject = SelectField("Subject", choices=[("", "-- Select Subject --")] + app.config['LEARNING_AREAS'], validators=[InputRequired()])
    class_name = SelectField("Class", choices=[("", "-- Select Class --")] + app.config['CLASS_LEVELS'], validators=[Optional()])

class AssessmentFilterForm(FlaskForm):
    subject = SelectField("Subject", choices=[("", "-- All Subjects --")] + app.config['LEARNING_AREAS'], validators=[Optional()])
    class_name = SelectField("Class", choices=[("", "-- All Classes --")] + app.config['CLASS_LEVELS'], validators=[Optional()])
    category = SelectField("Category", choices=[("", "-- All Categories --")] + app.config['ASSESSMENT_CATEGORIES'], validators=[Optional()])

class BulkImportForm(FlaskForm):
    excel_file = FileField("Excel File", validators=[
        InputRequired(),
        FileAllowed(['xlsx', 'xls'], 'Excel files only!')
    ])

class StudentBulkImportForm(FlaskForm):
    excel_file = FileField("Excel File", validators=[
        InputRequired(),
        FileAllowed(['xlsx', 'xls'], 'Excel files only!')
    ])

class QuestionBulkImportForm(FlaskForm):
    excel_file = FileField("Excel File", validators=[
        InputRequired(),
        FileAllowed(['xlsx', 'xls'], 'Excel files only!')
    ])

class SettingsForm(FlaskForm):
    current_term = SelectField("Current Term", choices=app.config['TERMS'], validators=[InputRequired()])
    current_academic_year = StringField("Current Academic Year", validators=[InputRequired()])
    current_session = StringField("Current Session", validators=[InputRequired()])
    assessment_active = BooleanField("Assessment Entry Active", default=True)


class QuestionForm(FlaskForm):
    question_text = TextAreaField("Question Text", validators=[InputRequired(), Length(min=10, max=1000)])
    question_type = SelectField("Question Type", choices=[
        ('mcq', 'Multiple Choice Question'),
        ('true_false', 'True/False'),
        ('short_answer', 'Short Answer')
    ], validators=[InputRequired()])
    options = TextAreaField("Options (for MCQ only)", validators=[Optional()], 
                          render_kw={"placeholder": "Enter options one per line (A, B, C, D)"})
    correct_answer = StringField("Correct Answer", validators=[InputRequired()], 
                               render_kw={"placeholder": "For MCQ: A, B, C, or D. For True/False: True or False"})
    difficulty = SelectField("Difficulty", choices=[
        ('easy', 'Easy'),
        ('medium', 'Medium'),
        ('hard', 'Hard')
    ], validators=[InputRequired()])
    explanation = TextAreaField("Explanation (Optional)", validators=[Optional(), Length(max=500)])


class QuizForm(FlaskForm):
    title = StringField("Quiz Title", validators=[InputRequired(), Length(min=3, max=200)])
    subject = SelectField("Subject", validators=[InputRequired()])
    description = TextAreaField("Description", validators=[Optional(), Length(max=500)])
    questions = SelectMultipleField("Questions", validators=[InputRequired()], 
                                   render_kw={"size": 10})
    time_limit = FloatField("Time Limit (minutes)", validators=[Optional(), NumberRange(min=1, max=180)])
    is_active = BooleanField("Active", default=True)


# -------------------------
# Login manager
# -------------------------
@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))

# -------------------------
# Decorators
# -------------------------
def admin_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if not current_user.is_authenticated:
            return login_manager.unauthorized()
        if not current_user.is_admin():
            abort(403)
        return f(*args, **kwargs)
    return wrapped

def teacher_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_teacher():
            abort(403)
        return f(*args, **kwargs)
    return decorated_function

def student_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_student():
            abort(403)
        return f(*args, **kwargs)
    return decorated_function

def student_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_student():
            abort(403)
        return f(*args, **kwargs)
    return decorated_function

# -------------------------
# Authentication Routes
# -------------------------
@app.route("/login", methods=["GET", "POST"])
def login():
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data.strip()).first()
        if user and user.check_password(form.password.data, bcrypt):
            login_user(user)
            log_activity(user, "login", f"User {user.username} logged in")
            flash("Logged in successfully", "success")
            next_page = request.args.get("next")
            return redirect(next_page or url_for("dashboard"))
        flash("Invalid credentials", "danger")
    return render_template("login.html", form=form)

@app.route("/logout")
@login_required
def logout():
    logout_user()
    flash("Logged out successfully", "info")
    return redirect(url_for("login"))

@app.route("/student/login", methods=["GET", "POST"])
def student_login():
    """Student login using first name and student number"""
    if current_user.is_authenticated:
        return redirect(url_for("dashboard"))
    
    form = StudentLoginForm()
    if form.validate_on_submit():
        first_name = form.first_name.data.strip()
        student_number = form.student_number.data.strip()
        
        # Find student by first name and student number
        student = Student.query.filter_by(first_name=first_name, student_number=student_number).first()
        
        if student:
            # Check if there's a user account for this student
            user = User.query.filter_by(username=student_number).first()
            if not user:
                # Create a student user account if it doesn't exist
                password = app.config['DEFAULT_STUDENT_PASSWORD']
                pw_hash = bcrypt.generate_password_hash(password).decode("utf-8")
                user = User(
                    username=student_number,
                    password_hash=pw_hash,
                    role="student"
                )
                db.session.add(user)
                db.session.commit()
            
            login_user(user)
            log_activity(user, "student_login", f"Student {student.full_name()} ({student.student_number}) logged in")
            flash("Student login successful", "success")
            return redirect(url_for("student_dashboard"))
        else:
            flash("Invalid first name or student number. Please check your details.", "danger")
    
    return render_template("student_login.html", form=form)

@app.route("/student/logout")
@login_required
def student_logout():
    """Logout student"""
    logout_user()
    flash("Logged out successfully", "info")
    return redirect(url_for("student_login"))

# -------------------------
# Dashboard Routes
# -------------------------
@app.route("/")
@login_required
def dashboard():
    if hasattr(current_user, 'is_student') and current_user.is_student():
        return redirect(url_for("student_dashboard"))
    
    # Teacher/Admin dashboard
    student_count = Student.query.count()
    assessment_count = Assessment.query.filter_by(archived=False).count()
    users_count = User.query.count()
    
    # For teachers, show only their assessments
    if hasattr(current_user, 'is_teacher') and current_user.is_teacher():
        recent = Assessment.query.filter_by(teacher_id=current_user.id, archived=False)\
            .order_by(Assessment.date_recorded.desc()).limit(8).all()
    else:
        recent = Assessment.query.filter_by(archived=False)\
            .order_by(Assessment.date_recorded.desc()).limit(8).all()
    
    return render_template(
        "dashboard.html",
        student_count=student_count,
        assessment_count=assessment_count,
        users_count=users_count,
        recent=recent
    )

@app.route("/student/dashboard")
@login_required
@student_required
def student_dashboard():
    """Student dashboard showing their assessments"""
    # Get student info using student number (which is the username)
    student = Student.query.filter_by(student_number=current_user.username).first()
    if not student:
        flash("Student record not found", "danger")
        return redirect(url_for("student_logout"))
    
    # Get filter parameters
    subject = request.args.get("subject", "")
    class_filter = request.args.get("class", "")
    
    # Get assessments
    query = Assessment.query.filter_by(student_id=student.id, archived=False)
    
    if subject:
        query = query.filter_by(subject=subject)
    if class_filter:
        query = query.filter_by(class_name=class_filter)
    
    assessments = query.order_by(Assessment.date_recorded.desc()).all()
    
    # Get unique subjects and classes for filter dropdowns
    subjects = sorted(set([a.subject for a in student.assessments if a.subject]))
    classes = sorted(set([a.class_name for a in student.assessments if a.class_name]))
    
    # Calculate summary
    summary = student.get_assessment_summary()
    final_percent = student.calculate_final_grade()
    gpa_grade = student.get_gpa_and_grade()
    
    # Calculate comment based on GPA
    def get_comment(gpa_str):
        try:
            gpa = float(gpa_str)
            if gpa == 4.0: return "Excellent"
            elif gpa == 3.5: return "Very Good"
            elif gpa == 3.0: return "Good"
            elif gpa == 2.5: return "Average"
            elif gpa == 2.0: return "Below Average"
            elif gpa == 1.5: return "Credit"
            elif gpa == 1.0: return "Satisfactory"
            elif gpa == 0.5: return "Pass"
            else: return "Fail"
        except (ValueError, TypeError):
            return None
    
    comment = get_comment(gpa_grade['gpa']) if gpa_grade['gpa'] != 'N/A' else None
    
    return render_template(
        "student_dashboard.html",
        student=student,
        assessments=assessments,
        summary=summary,
        final_percent=final_percent,
        gpa_grade=gpa_grade,
        comment=comment,
        subjects=subjects,
        classes=classes,
        selected_subject=subject,
        selected_class=class_filter,
        category_labels=app.config['CATEGORY_LABELS']
    )

# -------------------------
# Student Management Routes
# -------------------------
@app.route("/students")
@login_required
def students():
    search = request.args.get("search", "").strip()
    
    if search:
        students = Student.query.filter(
            (Student.student_number.ilike(f"%{search}%")) |
            (Student.first_name.ilike(f"%{search}%")) |
            (Student.last_name.ilike(f"%{search}%")) |
            (Student.reference_number.ilike(f"%{search}%"))
        ).order_by(Student.last_name, Student.first_name).all()
    else:
        students = Student.query.order_by(Student.last_name, Student.first_name).all()
    
    return render_template("students.html", students=students)

@app.route("/students/new", methods=["GET", "POST"])
@login_required
def student_new():
    # Only teachers and admins can create students
    if not (current_user.is_teacher() or current_user.is_admin()):
        abort(403)
        
    form = StudentForm()
    if form.validate_on_submit():
        exists = Student.query.filter_by(student_number=form.student_number.data.strip()).first()
        if exists:
            flash("Student number already exists", "warning")
        else:
            # Generate reference number (STU + random 6 digits)
            reference_number = f"STU{random.randint(100000, 999999)}"
            
            student = Student(
                student_number=form.student_number.data.strip(),
                first_name=form.first_name.data.strip(),
                last_name=form.last_name.data.strip(),
                middle_name=form.middle_name.data.strip() if form.middle_name.data else None,
                class_name=form.class_name.data if form.class_name.data else None,
                study_area=form.study_area.data if form.study_area.data else None,
                reference_number=reference_number
            )
            db.session.add(student)
            db.session.commit()
            
            log_activity(current_user, "create_student", f"Created student {student.full_name()} ({student.student_number})")
            flash(f"Student {student.full_name()} added successfully. Reference Number: {reference_number}", "success")
            return redirect(url_for("students"))
    
    return render_template("student_form.html", form=form, student=None)

@app.route("/students/<int:student_id>/edit", methods=["GET", "POST"])
@login_required
def student_edit(student_id):
    # Only teachers and admins can edit students
    if not (current_user.is_teacher() or current_user.is_admin()):
        abort(403)
        
    student = Student.query.get_or_404(student_id)
    form = StudentForm(obj=student)
    
    if form.validate_on_submit():
        student.student_number = form.student_number.data.strip()
        student.first_name = form.first_name.data.strip()
        student.last_name = form.last_name.data.strip()
        student.middle_name = form.middle_name.data.strip() if form.middle_name.data else None
        student.class_name = form.class_name.data if form.class_name.data else None
        student.study_area = form.study_area.data if form.study_area.data else None
        db.session.commit()
        log_activity(current_user, "edit_student", f"Edited student {student.full_name()} ({student.student_number})")
        flash(f"Student {student.full_name()} updated successfully", "success")
        return redirect(url_for("students"))
    
    return render_template("student_form.html", form=form, student=student)

@app.route("/students/<int:student_id>/delete", methods=["POST"])
@login_required
@admin_required
def student_delete(student_id):
    student = Student.query.get_or_404(student_id)
    student_name = student.full_name()
    db.session.delete(student)
    db.session.commit()
    log_activity(current_user, "delete_student", f"Deleted student {student_name} ({student.student_number})")
    flash(f"Student {student_name} deleted successfully", "info")
    return redirect(url_for("students"))

@app.route("/students/<int:student_id>")
@login_required
def student_view(student_id):
    student = Student.query.get_or_404(student_id)
    
    subject = request.args.get('subject')
    
    # Filter assessments by subject if specified
    if subject:
        assessments = [a for a in student.assessments if a.subject == subject]
    else:
        # Filter assessments by subject/class if teacher
        if hasattr(current_user, 'is_teacher') and current_user.is_teacher() and current_user.subject:
            assessments = [a for a in student.assessments if a.subject == current_user.subject]
        else:
            assessments = student.assessments
    
    # Get assessment summary and final grade
    summary = student.get_assessment_summary(subject)
    final_percent = student.calculate_final_grade(subject=subject)
    
    # Get all subjects for this student
    all_subjects = sorted(set(a.subject for a in student.assessments))
    
    # Calculate letter grade and GPA
    def get_letter_grade(percent):
        if percent >= 80: return 'A1'
        elif percent >= 70: return 'B2'
        elif percent >= 65: return 'B3'
        elif percent >= 60: return 'C4'
        elif percent >= 55: return 'C5'
        elif percent >= 50: return 'C6'
        elif percent >= 45: return 'D7'
        elif percent >= 40: return 'E8'
        else: return 'F9'
    
    def get_gpa(percent):
        if percent >= 80: return 4.0
        elif percent >= 70: return 3.5
        elif percent >= 65: return 3.0
        elif percent >= 60: return 2.5
        elif percent >= 55: return 2.0
        elif percent >= 50: return 1.5
        elif percent >= 45: return 1.0
        elif percent >= 40: return 0.5
        else: return 0.0
    
    letter_grade = get_letter_grade(final_percent) if final_percent is not None else None
    gpa = get_gpa(final_percent) if final_percent is not None else None
    
    def get_comment(gpa):
        if gpa == 4.0: return "Excellent"
        elif gpa == 3.5: return "Very Good"
        elif gpa == 3.0: return "Good"
        elif gpa == 2.5: return "Average"
        elif gpa == 2.0: return "Below Average"
        elif gpa == 1.5: return "Credit"
        elif gpa == 1.0: return "Satisfactory"
        elif gpa == 0.5: return "Pass"
        else: return "Fail"
    
    comment = get_comment(gpa) if gpa is not None else None
    
    return render_template(
        "student_view.html",
        student=student,
        assessments=assessments,
        summary=summary,
        final_percent=final_percent,
        letter_grade=letter_grade,
        gpa=gpa,
        comment=comment,
        subject=subject,
        all_subjects=all_subjects,
        category_labels=app.config['CATEGORY_LABELS'],
        study_areas_dict=dict(app.config['STUDY_AREAS'])
    )

@app.route("/students/bulk-import", methods=["GET", "POST"])
@login_required
def student_bulk_import():
    """Bulk import students from Excel file"""
    # Only teachers and admins can bulk import students
    if not (current_user.is_teacher() or current_user.is_admin()):
        abort(403)
        
    form = StudentBulkImportForm()
    
    if form.validate_on_submit():
        file = form.excel_file.data
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        # Save uploaded file
        file.save(filepath)
        
        try:
            # Import students
            importer = StudentBulkImporter(filepath)
            students_data = importer.import_students()
            
            # Process and save students
            success_count = 0
            error_count = 0
            errors = []
            
            for data in students_data:
                try:
                    # Check if student already exists
                    exists = Student.query.filter_by(student_number=data['student_number']).first()
                    if exists:
                        errors.append(f"Student {data['student_number']} already exists")
                        error_count += 1
                        continue
                    
                    # Generate reference number
                    reference_number = f"STU{random.randint(100000, 999999)}"
                    
                    student = Student(
                        student_number=data['student_number'],
                        first_name=data['first_name'],
                        last_name=data['last_name'],
                        middle_name=data.get('middle_name'),
                        class_name=data.get('class_name'),
                        study_area=data.get('study_area'),
                        reference_number=reference_number
                    )
                    db.session.add(student)
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"Error importing {data.get('student_number', 'unknown')}: {str(e)}")
                    error_count += 1
            
            db.session.commit()
            
            # Clean up uploaded file
            os.remove(filepath)
            
            flash(f"Bulk import completed. {success_count} students imported successfully. {error_count} errors.", "success")
            if errors:
                flash("Errors: " + "; ".join(errors[:5]), "warning")  # Show first 5 errors
            
            return redirect(url_for("students"))
            
        except Exception as e:
            flash(f"Error importing file: {str(e)}", "danger")
            return redirect(url_for("student_bulk_import"))
    
    return render_template("student_bulk_import.html", form=form)


@app.route("/teacher/questions/bulk_import", methods=["GET", "POST"])
@login_required
@teacher_required
def bulk_import_questions():
    """Bulk import questions from Excel file"""
    form = QuestionBulkImportForm()
    
    if form.validate_on_submit():
        file = form.excel_file.data
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        # Save uploaded file
        file.save(filepath)
        
        try:
            # Import questions
            importer = QuestionBulkImporter(filepath)
            questions_data = importer.import_questions()
            
            # Process and save questions
            success_count = 0
            error_count = 0
            errors = []
            
            for data in questions_data:
                try:
                    # Create question
                    question = Question(
                        subject=current_user.subject,
                        question_text=data['question_text'],
                        question_type=data['question_type'],
                        options=data['options'],
                        correct_answer=data['correct_answer'],
                        difficulty=data['difficulty'],
                        explanation=data['explanation'],
                        created_by=current_user.id
                    )
                    db.session.add(question)
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"Error importing question '{data.get('question_text', 'unknown')[:50]}...': {str(e)}")
                    error_count += 1
            
            db.session.commit()
            
            # Clean up uploaded file
            os.remove(filepath)
            
            flash(f"Bulk import completed. {success_count} questions imported successfully. {error_count} errors.", "success")
            if errors:
                flash("Errors: " + "; ".join(errors[:5]), "warning")  # Show first 5 errors
            
            return redirect(url_for("teacher_question_bank"))
            
        except Exception as e:
            flash(f"Error importing file: {str(e)}", "danger")
            return redirect(url_for("bulk_import_questions"))
    
    return render_template("question_bulk_import.html", form=form)


# -------------------------
# Assessment Routes
# -------------------------
@app.route("/assessments")
@login_required
def assessments_list():
    page = request.args.get("page", 1, type=int)
    subject = request.args.get("subject", "")
    class_name = request.args.get("class", "")
    category = request.args.get("category", "")
    
    per_page = app.config['ASSESSMENTS_PER_PAGE']
    
    # Build query based on user role and filters
    if hasattr(current_user, 'is_teacher') and current_user.is_teacher():
        query = Assessment.query.filter_by(teacher_id=current_user.id, archived=False)
    else:
        query = Assessment.query.filter_by(archived=False)
    
    if subject:
        query = query.filter_by(subject=subject)
    if class_name:
        query = query.filter_by(class_name=class_name)
    if category:
        query = query.filter_by(category=category)
    
    pagination = query.order_by(Assessment.date_recorded.desc())\
        .paginate(page=page, per_page=per_page, error_out=False)
    
    form = AssessmentFilterForm()
    form.subject.data = subject
    form.class_name.data = class_name
    form.category.data = category
    
    return render_template(
        "assessments.html",
        assessments=pagination.items,
        form=form,
        page=page,
        per_page=per_page,
        total=pagination.total,
        category_labels=app.config['CATEGORY_LABELS'],
        pagination=pagination
    )

@app.route("/assessments/new", methods=["GET", "POST"])
@login_required
def new_assessment():
    # Only teachers and admins can create assessments
    if not (current_user.is_teacher() or current_user.is_admin()):
        abort(403)
        
    form = AssessmentForm()
    
    # Populate student choices
    students = Student.query.all()
    form.student_name.choices = [("", "-- Select Student --")] + [(s.student_number, s.full_name()) for s in students]
    student_dict = {s.student_number: {'name': s.full_name(), 'ref': s.reference_number or ''} for s in students}
    
    # Get global settings
    settings = Setting.query.first()
    
    # Auto-fill subject and class for teachers
    if current_user.is_teacher() and current_user.subject:
        form.subject.data = current_user.subject
    if current_user.is_teacher() and current_user.class_name:
        form.class_name.data = current_user.class_name
    
    # Auto-fill global settings
    if settings:
        form.term.data = settings.current_term
        form.academic_year.data = settings.current_academic_year
        form.session.data = settings.current_session
    
    if form.validate_on_submit():
        # Get student_number from either dropdown or manual input
        student_number = form.student_name.data or form.student_number.data.strip()
        student = Student.query.filter_by(student_number=student_number).first()
        
        if not student:
            flash("Student not found. Please create the student first.", "danger")
        else:
            # Check if assessment already exists for this student, category, subject, term, academic_year, session
            existing_assessment = Assessment.query.filter_by(
                student_id=student.id,
                category=form.category.data,
                subject=form.subject.data,
                term=form.term.data,
                academic_year=form.academic_year.data,
                session=form.session.data
            ).first()
            
            if existing_assessment:
                flash(f"An assessment for {form.category.data} in {form.subject.data} already exists for this student in the same term, academic year, and session. Please update the existing assessment instead.", "warning")
                return redirect(url_for('student_view', student_id=student.id))
            
            # Set max_score based on category
            category = form.category.data
            max_score = app.config['CATEGORY_MAX_SCORES'].get(category, 100.0)
            
            # Validate score doesn't exceed max_score
            if form.score.data > max_score:
                flash(f"Score cannot exceed max score of {max_score}", "danger")
                return redirect(url_for('new_assessment'))
            
            # Auto-assign class from student if not specified
            class_name = form.class_name.data or student.class_name
            
            assessment = Assessment(
                student=student,
                category=category,
                subject=form.subject.data,
                class_name=class_name,
                score=float(form.score.data),
                max_score=max_score,
                term=form.term.data,
                academic_year=form.academic_year.data,
                session=form.session.data,
                assessor=form.assessor.data or current_user.username,
                teacher_id=current_user.id if hasattr(current_user, 'is_teacher') and current_user.is_teacher() else None,
                comments=form.comments.data
            )
            db.session.add(assessment)
            db.session.commit()
            log_activity(current_user, "create_assessment", f"Created assessment for {student.full_name()} ({assessment.category} in {assessment.subject})")
            flash(f"Assessment saved for {student.full_name()}", "success")
            return redirect(url_for("student_view", student_id=student.id))
    
    return render_template("assessment_form.html", form=form, students=students, student_dict=student_dict)

@app.route("/assessments/<int:assessment_id>/edit", methods=["GET", "POST"])
@login_required
def assessment_edit(assessment_id):
    assessment = Assessment.query.get_or_404(assessment_id)
    
    # Only teachers and admins can edit assessments
    # Teachers can only edit their own assessments
    if not (current_user.is_admin() or 
            (current_user.is_teacher() and assessment.teacher_id == current_user.id)):
        abort(403)
        
    form = AssessmentForm(obj=assessment)
    
    # Populate student choices
    students = Student.query.all()
    form.student_name.choices = [("", "-- Select Student --")] + [(s.student_number, s.full_name()) for s in students]
    student_dict = {s.student_number: {'name': s.full_name(), 'ref': s.reference_number or ''} for s in students}
    
    # Pre-fill form
    form.student_name.data = assessment.student.student_number
    form.student_number.data = assessment.student.student_number
    form.reference_number.data = assessment.student.reference_number
    
    if form.validate_on_submit():
        # Set max_score based on category
        category = form.category.data
        max_score = app.config['CATEGORY_MAX_SCORES'].get(category, 100.0)
        
        # Validate score doesn't exceed max_score
        if form.score.data > max_score:
            flash(f"Score cannot exceed max score of {max_score}", "danger")
            return redirect(url_for('assessment_edit', assessment_id=assessment_id))
        
        assessment.category = category
        assessment.subject = form.subject.data
        assessment.class_name = form.class_name.data
        assessment.score = float(form.score.data)
        assessment.max_score = max_score
        assessment.term = form.term.data
        assessment.academic_year = form.academic_year.data
        assessment.session = form.session.data
        assessment.assessor = form.assessor.data
        assessment.comments = form.comments.data
        db.session.commit()
        log_activity(current_user, "edit_assessment", f"Edited assessment for {assessment.student.full_name()} ({assessment.category} in {assessment.subject})")
        flash("Assessment updated successfully", "success")
        return redirect(url_for("student_view", student_id=assessment.student_id))
    
    return render_template("assessment_form.html", form=form, assessment=assessment, students=students, student_dict=student_dict)

@app.route("/assessments/<int:assessment_id>/delete", methods=["POST"])
@login_required
def assessment_delete(assessment_id):
    assessment = Assessment.query.get_or_404(assessment_id)
    
    # Only teachers and admins can delete assessments
    # Teachers can only delete their own assessments
    if not (current_user.is_admin() or 
            (current_user.is_teacher() and assessment.teacher_id == current_user.id)):
        abort(403)
        
    student_id = assessment.student_id
    db.session.delete(assessment)
    db.session.commit()
    log_activity(current_user, "delete_assessment", f"Deleted assessment for {assessment.student.full_name()} ({assessment.category} in {assessment.subject})")
    flash("Assessment deleted successfully", "info")
    return redirect(url_for("student_view", student_id=student_id))

@app.route("/assessments/<int:assessment_id>/archive", methods=["POST"])
@login_required
def assessment_archive(assessment_id):
    assessment = Assessment.query.get_or_404(assessment_id)
    
    # Only teachers and admins can archive assessments
    # Teachers can only archive their own assessments
    if not (current_user.is_admin() or 
            (current_user.is_teacher() and assessment.teacher_id == current_user.id)):
        abort(403)
        
    assessment.archived = True
    db.session.commit()
    flash("Assessment archived successfully", "info")
    return redirect(request.referrer or url_for("assessments"))

@app.route("/assessments/<int:assessment_id>/unarchive", methods=["POST"])
@login_required
def assessment_unarchive(assessment_id):
    assessment = Assessment.query.get_or_404(assessment_id)
    
    # Only teachers and admins can unarchive assessments
    # Teachers can only unarchive their own assessments
    if not (current_user.is_admin() or 
            (current_user.is_teacher() and assessment.teacher_id == current_user.id)):
        abort(403)
        
    assessment.archived = False
    db.session.commit()
    flash("Assessment unarchived successfully", "info")
    return redirect(request.referrer or url_for("assessments"))

@app.route("/assessments/archived")
@login_required
def assessments_archived():
    page = request.args.get('page', 1, type=int)
    subject = request.args.get('subject', '')
    class_name = request.args.get('class', '')
    category = request.args.get('category', '')
    
    per_page = app.config['ASSESSMENTS_PER_PAGE']
    
    # Build query based on user role and filters - only archived
    if hasattr(current_user, 'is_teacher') and current_user.is_teacher():
        query = Assessment.query.filter_by(teacher_id=current_user.id, archived=True)
    else:
        query = Assessment.query.filter_by(archived=True)
    
    if subject:
        query = query.filter_by(subject=subject)
    if class_name:
        query = query.filter_by(class_name=class_name)
    if category:
        query = query.filter_by(category=category)
    
    pagination = query.order_by(Assessment.date_recorded.desc())\
        .paginate(page=page, per_page=per_page, error_out=False)
    
    form = AssessmentFilterForm()
    form.subject.data = subject
    form.class_name.data = class_name
    form.category.data = category
    
    return render_template(
        "assessments.html",
        assessments=pagination.items,
        pagination=pagination,
        form=form,
        category_labels=app.config['CATEGORY_LABELS'],
        archived=True
    )

# -------------------------
@app.route("/users")
@login_required
@admin_required
def users():
    teachers_admins = User.query.filter(User.role.in_(['admin', 'teacher'])).order_by(User.username).all()
    students = User.query.filter_by(role='student').order_by(User.username).all()
    return render_template("users.html", teachers_admins=teachers_admins, students=students)

@app.route("/users/new", methods=["GET", "POST"])
@login_required
@admin_required
def create_user():
    form = UserForm()
    
    if form.validate_on_submit():
        if User.query.filter_by(username=form.username.data.strip()).first():
            flash("Username already exists", "warning")
        else:
            pw_hash = bcrypt.generate_password_hash(form.password.data).decode("utf-8")
            user = User(
                username=form.username.data.strip(),
                password_hash=pw_hash,
                role=form.role.data,
                subject=form.subject.data if form.subject.data else None,
                class_name=form.class_name.data if form.class_name.data else None
            )
            db.session.add(user)
            db.session.commit()
            log_activity(current_user, "create_user", f"Created user {user.username} with role {user.role}")
            flash(f"User {user.username} created successfully", "success")
            return redirect(url_for("users"))
    
    return render_template("user_form.html", form=form)

@app.route("/users/<int:user_id>/edit", methods=["GET", "POST"])
@login_required
@admin_required
def edit_user(user_id):
    user = User.query.get_or_404(user_id)
    form = EditUserForm(role=user.role)
    
    if form.validate_on_submit():
        user.role = form.role.data
        user.subject = form.subject.data if form.subject.data else None
        user.class_name = form.class_name.data if form.class_name.data else None
        db.session.commit()
        log_activity(current_user, "edit_user", f"Edited user {user.username}")
        flash(f"User {user.username} updated successfully", "success")
        return redirect(url_for("users"))
    
    # Pre-fill form
    if user.subject:
        form.subject.data = user.subject
    if user.class_name:
        form.class_name.data = user.class_name
    
    return render_template("edit_user.html", form=form, user=user)

@app.route("/users/<int:user_id>/reset_password", methods=["GET", "POST"])
@login_required
@admin_required
def reset_password(user_id):
    user = User.query.get_or_404(user_id)
    form = PasswordResetForm()
    
    if form.validate_on_submit():
        user.password_hash = bcrypt.generate_password_hash(form.password.data).decode("utf-8")
        db.session.commit()
        log_activity(current_user, "reset_password", f"Reset password for user {user.username}")
        flash(f"Password reset successfully for {user.username}", "success")
        return redirect(url_for("users"))
    
    return render_template("reset_password.html", form=form, user=user)

@app.route("/users/<int:user_id>/delete", methods=["POST"])
@login_required
@admin_required
def delete_user(user_id):
    if current_user.id == user_id:
        flash("You cannot delete your own account", "danger")
        return redirect(url_for("users"))
    
    user = User.query.get_or_404(user_id)
    username = user.username
    db.session.delete(user)
    db.session.commit()
    log_activity(current_user, "delete_user", f"Deleted user {username}")
    flash(f"User {username} deleted successfully", "info")
    return redirect(url_for("users"))

# -------------------------
# Admin Settings Routes
# -------------------------
@app.route("/admin/settings", methods=["GET", "POST"])
@login_required
@admin_required
def admin_settings():
    """Admin can configure global settings"""
    settings = Setting.query.first()
    if not settings:
        settings = Setting()
        db.session.add(settings)
        db.session.commit()
    
    form = SettingsForm(obj=settings)
    
    if form.validate_on_submit():
        settings.current_term = form.current_term.data
        settings.current_academic_year = form.current_academic_year.data
        settings.current_session = form.current_session.data
        settings.assessment_active = form.assessment_active.data
        db.session.commit()
        flash("Settings updated successfully", "success")
        return redirect(url_for("admin_settings"))
    
    return render_template("admin_settings.html", form=form, settings=settings)

@app.route("/admin/activity-logs")
@login_required
@admin_required
def admin_activity_logs():
    """Admin can view activity logs"""
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    logs = ActivityLog.query.order_by(ActivityLog.timestamp.desc()).paginate(page=page, per_page=per_page, error_out=False)
    
    return render_template("activity_logs.html", logs=logs)

# -------------------------
# Teacher Routes
# -------------------------
@app.route("/users/<int:user_id>/assign-subject", methods=["GET", "POST"])
@login_required
@admin_required
def assign_teacher_subject(user_id):
    """Admin can assign subject specialization to teachers"""
    user = User.query.get_or_404(user_id)
    if not user.is_teacher():
        flash("This user is not a teacher", "danger")
        return redirect(url_for("users"))
        
    form = TeacherAssignmentForm()
    
    if form.validate_on_submit():
        user.subject = form.subject.data
        user.class_name = form.class_name.data if form.class_name.data else None
        db.session.commit()
        flash(f"Subject assigned to {user.username}: {dict(app.config['LEARNING_AREAS']).get(form.subject.data)}", "success")
        return redirect(url_for("users"))
    
    if user.subject:
        form.subject.data = user.subject
    if user.class_name:
        form.class_name.data = user.class_name
    
    return render_template("teacher_subject.html", form=form, teacher=user)

@app.route("/teacher/subject", methods=["GET", "POST"])
@login_required
@teacher_required
def teacher_subject():
    """Teacher can set their subject specialization"""
    user = current_user
    
    form = TeacherAssignmentForm()
    
    if form.validate_on_submit():
        user.subject = form.subject.data
        user.class_name = form.class_name.data if form.class_name.data else None
        db.session.commit()
        flash(f"Subject updated: {dict(app.config['LEARNING_AREAS']).get(form.subject.data)}", "success")
        return redirect(url_for("dashboard"))
    
    if user.subject:
        form.subject.data = user.subject
    if user.class_name:
        form.class_name.data = user.class_name
    
    return render_template("teacher_subject.html", form=form, teacher=None)


# -------------------------------
# Question Bank Routes
# -------------------------------

@app.route("/teacher/question-bank")
@login_required
def teacher_question_bank():
    """Teacher can view and manage their subject questions, Admin can view all"""
    if not (current_user.is_teacher() or current_user.is_admin()):
        abort(403)
    
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    # Admin can see all questions, teachers see their subject
    if current_user.is_admin():
        query = Question.query
        # Allow admin to filter by subject
        subject_filter = request.args.get('subject')
        if subject_filter:
            query = query.filter_by(subject=subject_filter)
    else:
        query = Question.query.filter_by(subject=current_user.subject)
    
    # Filter by status if specified
    status_filter = request.args.get('status')
    if status_filter:
        query = query.filter_by(status=status_filter)
    
    questions = query.order_by(Question.created_at.desc()).paginate(page=page, per_page=per_page)
    
    # Get subjects for admin filter
    subjects = []
    if current_user.is_admin():
        subjects = db.session.query(Question.subject).distinct().all()
        subjects = [s[0] for s in subjects]
    
    return render_template("teacher_question_bank.html", questions=questions, 
                         status_filter=status_filter, subject_filter=request.args.get('subject'), 
                         subjects=subjects, is_admin=current_user.is_admin())


@app.route("/teacher/questions/new", methods=["GET", "POST"])
@login_required
@teacher_required
def create_question():
    """Teacher can create new questions"""
    form = QuestionForm()
    
    if form.validate_on_submit():
        question = Question(
            subject=current_user.subject,
            question_text=form.question_text.data,
            question_type=form.question_type.data,
            options=form.options.data if form.options.data else None,
            correct_answer=form.correct_answer.data,
            difficulty=form.difficulty.data,
            explanation=form.explanation.data,
            created_by=current_user.id
        )
        db.session.add(question)
        db.session.commit()
        
        # Log activity
        log_activity(current_user, "create_question", f"Created question ID {question.id} for {question.subject}")
        
        flash("Question created successfully and submitted for approval", "success")
        return redirect(url_for("teacher_question_bank"))
    
    return render_template("question_form.html", form=form, title="Create Question")


@app.route("/teacher/questions/<int:question_id>/edit", methods=["GET", "POST"])
@login_required
@teacher_required
def edit_question(question_id):
    """Teacher can edit their pending questions"""
    question = Question.query.get_or_404(question_id)
    
    # Check permissions
    if not question.can_edit(current_user):
        abort(403)
    
    form = QuestionForm(obj=question)
    
    if form.validate_on_submit():
        question.question_text = form.question_text.data
        question.question_type = form.question_type.data
        question.options = form.options.data if form.options.data else None
        question.correct_answer = form.correct_answer.data
        question.difficulty = form.difficulty.data
        question.explanation = form.explanation.data
        question.updated_at = datetime.utcnow()
        
        db.session.commit()
        
        # Log activity
        log_activity(current_user, "edit_question", f"Edited question ID {question.id}")
        
        flash("Question updated successfully", "success")
        return redirect(url_for("teacher_question_bank"))
    
    return render_template("question_form.html", form=form, title="Edit Question", question=question)


@app.route("/teacher/questions/<int:question_id>/delete", methods=["POST"])
@login_required
@teacher_required
def delete_question(question_id):
    """Teacher can delete their pending questions"""
    question = Question.query.get_or_404(question_id)
    
    # Check permissions
    if not question.can_edit(current_user):
        abort(403)
    
    db.session.delete(question)
    db.session.commit()
    
    # Log activity
    log_activity(current_user, "delete_question", f"Deleted question ID {question.id}")
    
    flash("Question deleted successfully", "success")
    return redirect(url_for("teacher_question_bank"))


@app.route("/admin/question-bank")
@login_required
@admin_required
def admin_question_bank():
    """Admin can moderate all questions"""
    page = request.args.get('page', 1, type=int)
    per_page = 20
    
    # Get all questions
    query = Question.query
    
    # Filter by status if specified
    status_filter = request.args.get('status', 'pending')
    if status_filter:
        query = query.filter_by(status=status_filter)
    
    # Filter by subject if specified
    subject_filter = request.args.get('subject')
    if subject_filter:
        query = query.filter_by(subject=subject_filter)
    
    questions = query.order_by(Question.created_at.desc()).paginate(page=page, per_page=per_page)
    
    # Get all subjects for filter
    subjects = db.session.query(Question.subject).distinct().all()
    subjects = [s[0] for s in subjects]
    
    return render_template("admin_question_bank.html", questions=questions, 
                         status_filter=status_filter, subject_filter=subject_filter, subjects=subjects)


@app.route("/admin/questions/<int:question_id>/approve", methods=["POST"])
@login_required
@admin_required
def approve_question(question_id):
    """Admin can approve questions"""
    question = Question.query.get_or_404(question_id)
    
    action = request.form.get('action')
    if action == 'approve':
        question.status = 'approved'
        question.approved_by = current_user.id
        flash("Question approved successfully", "success")
    elif action == 'reject':
        question.status = 'rejected'
        question.approved_by = current_user.id
        question.rejection_reason = request.form.get('rejection_reason')
        flash("Question rejected", "warning")
    
    db.session.commit()
    
    # Log activity
    log_activity(current_user, "moderate_question", f"{action}d question ID {question.id}")
    
    return redirect(url_for("admin_question_bank"))


@app.route("/teacher/questions/<int:question_id>/approve", methods=["POST"])
@login_required
@teacher_required
def teacher_approve_question(question_id):
    """Teacher can approve questions in their subject"""
    question = Question.query.get_or_404(question_id)
    
    # Check if question is in teacher's subject
    if question.subject != current_user.subject:
        abort(403)
    
    action = request.form.get('action')
    if action == 'approve':
        question.status = 'approved'
        question.approved_by = current_user.id
        flash("Question approved successfully", "success")
    elif action == 'reject':
        question.status = 'rejected'
        question.approved_by = current_user.id
        question.rejection_reason = request.form.get('rejection_reason')
        flash("Question rejected", "warning")
    
    db.session.commit()
    
    # Log activity
    log_activity(current_user, "moderate_question", f"{action}d question ID {question.id}")
    
    return redirect(url_for("teacher_question_bank"))


@app.route("/student/questions")
@login_required
@student_required
def student_questions():
    """Student can view and answer questions"""
    page = request.args.get('page', 1, type=int)
    per_page = 10
    
    # Get approved questions for student's subjects
    # For now, get questions from all subjects, but in production this should be filtered
    # based on student's enrolled subjects
    questions = Question.query.filter_by(status='approved').order_by(Question.created_at.desc()).paginate(page=page, per_page=per_page)
    
    # Get student's previous attempts
    attempts = {attempt.question_id: attempt for attempt in 
               QuestionAttempt.query.filter_by(student_id=current_user.id).all()}
    
    return render_template("student_questions.html", questions=questions, attempts=attempts)


@app.route("/student/questions/<int:question_id>/attempt", methods=["POST"])
@login_required
@student_required
def attempt_question(question_id):
    """Student submits answer to a question"""
    question = Question.query.get_or_404(question_id)
    
    if question.status != 'approved':
        abort(404)
    
    student_answer = request.form.get('answer')
    if not student_answer:
        flash("Please provide an answer", "danger")
        return redirect(url_for("student_questions"))
    
    # Check if correct
    is_correct = False
    if question.question_type == 'mcq':
        is_correct = student_answer.strip().upper() == question.correct_answer.strip().upper()
    elif question.question_type == 'true_false':
        is_correct = student_answer.lower() == question.correct_answer.lower()
    else:  # short_answer - for now, simple string match, but could be more sophisticated
        is_correct = student_answer.strip().lower() == question.correct_answer.strip().lower()
    
    # Record attempt
    attempt = QuestionAttempt(
        student_id=current_user.id,
        question_id=question_id,
        student_answer=student_answer,
        is_correct=is_correct
    )
    db.session.add(attempt)
    db.session.commit()
    
    # Log activity
    log_activity(current_user, "attempt_question", f"Answered question ID {question.id}, correct: {is_correct}")
    
    if is_correct:
        flash("Correct answer!", "success")
    else:
        flash(f"Incorrect. The correct answer is: {question.correct_answer}", "warning")
    
    return redirect(url_for("student_questions"))


@app.route("/teacher/quizzes")
@login_required
def teacher_quizzes():
    """Teacher can view and manage quizzes for their subject, Admin can view all"""
    if not (current_user.is_teacher() or current_user.is_admin()):
        abort(403)
    
    if current_user.is_admin():
        quizzes = Quiz.query.order_by(Quiz.created_at.desc()).all()
    else:
        # Teachers see quizzes for their subject
        quizzes = Quiz.query.filter_by(subject=current_user.subject).order_by(Quiz.created_at.desc()).all()
    
    return render_template("teacher_quizzes.html", quizzes=quizzes)


@app.route("/teacher/quizzes/new", methods=["GET", "POST"])
@login_required
def create_quiz():
    """Teacher can create new quizzes for their subject, Admin can create for any subject"""
    if not (current_user.is_teacher() or current_user.is_admin()):
        abort(403)
    
    form = QuizForm()
    
    # Set subject choices based on user role
    if current_user.is_admin():
        # Admin can choose any subject
        form.subject.choices = [(subject, subject.replace('_', ' ').title()) for subject in app.config['LEARNING_AREAS']]
    else:
        # Teachers are limited to their subject
        form.subject.choices = [(current_user.subject, current_user.subject.replace('_', ' ').title())]
        form.subject.data = current_user.subject
    
    if form.validate_on_submit():
        # Get approved questions for the selected subject
        questions = Question.query.filter_by(subject=form.subject.data, status='approved').all()
        selected_question_ids = [int(q) for q in form.questions.data if q.isdigit()]
        
        # Validate that selected questions exist and are approved
        valid_questions = [q for q in questions if q.id in selected_question_ids]
        
        quiz = Quiz(
            title=form.title.data,
            subject=form.subject.data,
            description=form.description.data,
            questions=[q.id for q in valid_questions],
            time_limit=form.time_limit.data,
            created_by=current_user.id
        )
        db.session.add(quiz)
        db.session.commit()
        
        # Log activity
        log_activity(current_user, "create_quiz", f"Created quiz '{quiz.title}' with {len(quiz.questions)} questions")
        
        flash("Quiz created successfully", "success")
        return redirect(url_for("teacher_quizzes"))
    
    # For GET request, populate questions based on subject
    subject = request.args.get('subject', current_user.subject if current_user.is_teacher() else None)
    if subject:
        questions = Question.query.filter_by(subject=subject, status='approved').all()
        form.subject.data = subject
    else:
        questions = []
    
    form.questions.choices = [(str(q.id), f"{q.question_text[:50]}{'...' if len(q.question_text) > 50 else ''} ({q.difficulty.title()}, {q.question_type.upper()})") for q in questions]
    
    return render_template("quiz_form.html", form=form, available_questions=questions, quiz=None)


@app.route("/student/quizzes")
@login_required
@student_required
def student_quizzes():
    """Student can view available quizzes"""
    # For now, show all active quizzes, but should filter by student's subjects
    quizzes = Quiz.query.filter_by(is_active=True).order_by(Quiz.created_at.desc()).all()
    
    # Get student's previous attempts
    attempts = {attempt.quiz_id: attempt for attempt in 
               QuizAttempt.query.filter_by(student_id=current_user.id).all()}
    
    return render_template("student_quizzes.html", quizzes=quizzes, attempts=attempts)


@app.route("/student/quizzes/<int:quiz_id>/take", methods=["GET", "POST"])
@login_required
@student_required
def take_quiz(quiz_id):
    """Student takes a quiz"""
    quiz = Quiz.query.get_or_404(quiz_id)
    
    if not quiz.is_active:
        abort(404)
    
    # Check if student already attempted this quiz
    existing_attempt = QuizAttempt.query.filter_by(student_id=current_user.id, quiz_id=quiz_id).first()
    if existing_attempt:
        flash("You have already taken this quiz", "warning")
        return redirect(url_for("student_quizzes"))
    
    questions = Question.query.filter(Question.id.in_(quiz.questions)).all()
    questions_dict = {q.id: q for q in questions}
    
    if request.method == 'POST':
        # Process quiz submission
        answers = {}
        correct_count = 0
        question_results = {}
        
        for qid in quiz.questions:
            answer = request.form.get(f'answer_{qid}')
            if answer:
                question = questions_dict.get(int(qid))
                if question:
                    is_correct = False
                    if question.question_type == 'mcq':
                        is_correct = answer.strip().upper() == question.correct_answer.strip().upper()
                    elif question.question_type == 'true_false':
                        is_correct = answer.lower() == question.correct_answer.lower()
                    else:
                        is_correct = answer.strip().lower() == question.correct_answer.strip().lower()
                    
                    if is_correct:
                        correct_count += 1
                    
                    # Store question result for display
                    question_results[qid] = {
                        'student_answer': answer,
                        'is_correct': is_correct,
                        'correct_answer': question.correct_answer
                    }
                    
                    # Record individual question attempt
                    attempt = QuestionAttempt(
                        student_id=current_user.id,
                        question_id=qid,
                        student_answer=answer,
                        is_correct=is_correct
                    )
                    db.session.add(attempt)
        
        # Record quiz attempt
        quiz_attempt = QuizAttempt(
            student_id=current_user.id,
            quiz_id=quiz_id,
            score=correct_count,
            total_questions=len(quiz.questions),
            correct_answers=correct_count,
            completed_at=datetime.utcnow()
        )
        db.session.add(quiz_attempt)
        db.session.commit()
        
        # Log activity
        log_activity(current_user, "complete_quiz", f"Completed quiz '{quiz.title}' with score {correct_count}/{len(quiz.questions)}")
        
        # Store quiz results temporarily in session (expires in 2 hours)
        import time
        session['quiz_results'] = {
            'quiz_id': quiz_id,
            'quiz_title': quiz.title,
            'score': correct_count,
            'total_questions': len(quiz.questions),
            'percentage': round((correct_count / len(quiz.questions)) * 100, 1) if len(quiz.questions) > 0 else 0,
            'completed_at': datetime.utcnow().timestamp(),
            'question_results': question_results  # Store individual question results
        }
        session.modified = True
        
        return redirect(url_for("quiz_results"))
    
    return render_template("take_quiz.html", quiz=quiz, questions=questions_dict)


@app.route("/quiz/results")
@login_required
@student_required
def quiz_results():
    """Display quiz results temporarily (for 2 hours)"""
    quiz_results = session.get('quiz_results')
    
    if not quiz_results:
        flash("No quiz results available", "warning")
        return redirect(url_for("student_quizzes"))
    
    # Check if results are still valid (within 2 hours)
    import time
    current_time = time.time()
    results_time = quiz_results.get('completed_at', 0)
    
    if current_time - results_time > 7200:  # 2 hours in seconds
        session.pop('quiz_results', None)
        flash("Quiz results have expired", "info")
        return redirect(url_for("student_quizzes"))
    
    # Get quiz and questions for detailed display
    quiz = Quiz.query.get_or_404(quiz_results['quiz_id'])
    questions = {}
    for q_id in quiz.questions:
        question = Question.query.get(q_id)
        if question:
            questions[q_id] = question
    
    return render_template("quiz_results.html", 
                         quiz_results=quiz_results, 
                         quiz=quiz, 
                         questions=questions)


@app.route("/teacher/quizzes/<int:quiz_id>")
@login_required
def quiz_detail(quiz_id):
    """View quiz details"""
    if not (current_user.is_teacher() or current_user.is_admin()):
        abort(403)
    
    quiz = Quiz.query.get_or_404(quiz_id)
    
    # Check permissions: admin can see all, teachers can see their subject
    if not current_user.is_admin() and quiz.subject != current_user.subject:
        abort(403)
    
    # Get questions for this quiz
    questions = {}
    for q_id in quiz.questions:
        question = Question.query.get(q_id)
        if question:
            questions[q_id] = question
    
    return render_template("quiz_detail.html", quiz=quiz, questions=questions)


@app.route("/teacher/quizzes/<int:quiz_id>/edit", methods=["GET", "POST"])
@login_required
def edit_quiz(quiz_id):
    """Edit existing quiz"""
    if not (current_user.is_teacher() or current_user.is_admin()):
        abort(403)
    
    quiz = Quiz.query.get_or_404(quiz_id)
    
    # Check permissions: admin can edit all, teachers can edit their subject quizzes
    if not current_user.is_admin() and quiz.subject != current_user.subject:
        abort(403)
    
    form = QuizForm()
    
    if form.validate_on_submit():
        quiz.title = form.title.data
        quiz.description = form.description.data
        quiz.subject = form.subject.data
        quiz.time_limit = form.time_limit.data if form.time_limit.data else None
        quiz.is_active = form.is_active.data
        
        # Handle question selection
        selected_questions = request.form.getlist('questions')
        quiz.questions = [int(q) for q in selected_questions if q.isdigit()]
        
        db.session.commit()
        log_activity(current_user, "edit_quiz", f"Edited quiz '{quiz.title}'")
        flash("Quiz updated successfully", "success")
        return redirect(url_for("teacher_quizzes"))
    
    # Pre-populate form
    form.title.data = quiz.title
    form.description = form.description.data if form.description.data else quiz.description
    form.subject.data = quiz.subject
    form.time_limit.data = quiz.time_limit
    form.is_active.data = quiz.is_active
    
    # Get available questions for this subject
    available_questions = Question.query.filter_by(
        subject=quiz.subject, 
        status='approved'
    ).all()
    
    return render_template("quiz_form.html", form=form, quiz=quiz, available_questions=available_questions)


@app.route("/teacher/quizzes/<int:quiz_id>/delete", methods=["POST"])
@login_required
def delete_quiz(quiz_id):
    """Delete a quiz"""
    if not (current_user.is_teacher() or current_user.is_admin()):
        abort(403)
    
    quiz = Quiz.query.get_or_404(quiz_id)
    
    # Check permissions: admin can delete all, teachers can delete their subject quizzes
    if not current_user.is_admin() and quiz.subject != current_user.subject:
        abort(403)
    
    quiz_title = quiz.title
    
    # Delete associated attempts
    QuizAttempt.query.filter_by(quiz_id=quiz_id).delete()
    
    # Delete the quiz
    db.session.delete(quiz)
    db.session.commit()
    
    log_activity(current_user, "delete_quiz", f"Deleted quiz '{quiz_title}'")
    flash(f"Quiz '{quiz_title}' deleted successfully", "success")
    return redirect(url_for("teacher_quizzes"))


@app.route("/teacher/quizzes/<int:quiz_id>/results")
@login_required
def quiz_results_view(quiz_id):
    """View results of a specific quiz"""
    if not (current_user.is_teacher() or current_user.is_admin()):
        abort(403)
    
    quiz = Quiz.query.get_or_404(quiz_id)
    
    # Check permissions: admin can see all, teachers can see their subject quizzes
    if not current_user.is_admin() and quiz.subject != current_user.subject:
        abort(403)
    
    # Get all attempts for this quiz
    attempts = QuizAttempt.query.filter_by(quiz_id=quiz_id).order_by(QuizAttempt.completed_at.desc()).all()
    
    # Get student details
    student_ids = [attempt.student_id for attempt in attempts]
    students = {student.id: student for student in Student.query.filter(Student.id.in_(student_ids)).all()}
    
    return render_template("quiz_results_view.html", quiz=quiz, attempts=attempts, students=students)


@app.route("/teacher/quiz-results")
@login_required
def teacher_quiz_results():
    """Teacher can view all quiz results for their subject, Admin can view all"""
    if not (current_user.is_teacher() or current_user.is_admin()):
        abort(403)
    
    # Get quizzes based on permissions
    if current_user.is_admin():
        quizzes = Quiz.query.order_by(Quiz.created_at.desc()).all()
    else:
        quizzes = Quiz.query.filter_by(subject=current_user.subject).order_by(Quiz.created_at.desc()).all()
    
    # Get attempts for these quizzes
    quiz_ids = [quiz.id for quiz in quizzes]
    attempts = QuizAttempt.query.filter(QuizAttempt.quiz_id.in_(quiz_ids)).order_by(QuizAttempt.completed_at.desc()).all()
    
    # Group attempts by quiz
    attempts_by_quiz = {}
    for attempt in attempts:
        if attempt.quiz_id not in attempts_by_quiz:
            attempts_by_quiz[attempt.quiz_id] = []
        attempts_by_quiz[attempt.quiz_id].append(attempt)
    
    # Get student details
    student_ids = list(set(attempt.student_id for attempt in attempts))
    students = {student.id: student for student in Student.query.filter(Student.id.in_(student_ids)).all()}
    
    return render_template("teacher_quiz_results.html", quizzes=quizzes, attempts_by_quiz=attempts_by_quiz, students=students)


@app.route("/admin/archive-term", methods=["POST"])
@login_required
@admin_required
def archive_term():
    """Archive assessments for the previous term"""
    settings = Setting.query.first()
    if not settings:
        flash("No settings found", "danger")
        return redirect(url_for("admin_settings"))
    
    # Archive assessments not in current term
    assessments = Assessment.query.filter(
        (Assessment.term != settings.current_term) |
        (Assessment.academic_year != settings.current_academic_year)
    ).filter_by(archived=False).all()
    
    for assessment in assessments:
        assessment.archived = True
    
    db.session.commit()
    flash(f"Archived {len(assessments)} assessments from previous terms", "success")
    return redirect(url_for("admin_settings"))

# -------------------------
# API Endpoints
# -------------------------
@app.route("/api/student_search")
@login_required
def student_search():
    query = request.args.get("q", "").strip()
    
    if not query:
        return jsonify({"results": []})
    
    matches = Student.query.filter(
        (Student.student_number.ilike(f"%{query}%")) |
        (Student.first_name.ilike(f"%{query}%")) |
        (Student.last_name.ilike(f"%{query}%"))
    ).limit(10).all()
    
    results = [
        {
            "student_number": student.student_number,
            "name": student.full_name(),
            "reference_number": student.reference_number
        }
        for student in matches
    ]
    
    return jsonify({"results": results})

@app.route("/api/teacher/assessments")
@login_required
@teacher_required
def teacher_assessments_api():
    """Get assessments for teacher's subject - DIFFERENT NAME to avoid conflict"""
    if not current_user.subject:
        return jsonify({"assessments": []})
    
    assessments = Assessment.query.filter_by(
        subject=current_user.subject,
        teacher_id=current_user.id
    ).order_by(Assessment.date_recorded.desc()).limit(50).all()
    
    result = []
    for a in assessments:
        result.append({
            "student_name": a.student.full_name(),
            "student_number": a.student.student_number,
            "category": a.category,
            "score": a.score,
            "max_score": a.max_score,
            "percentage": a.get_percentage(),
            "class_name": a.class_name,
            "date": a.date_recorded.strftime("%Y-%m-%d")
        })
    
    return jsonify({"assessments": result})

# -------------------------
# Export Routes
# -------------------------
@app.route("/export/csv")
@login_required
def export_csv():
    assessments = Assessment.query.filter_by(archived=False)\
        .order_by(Assessment.date_recorded.desc()).all()
    
    # Create CSV in memory
    si = io.StringIO()
    writer = csv.writer(si)
    
    # Write header
    writer.writerow([
        "student_number",
        "name",
        "category",
        "subject",
        "score",
        "max_score",
        "percentage",
        "term",
        "academic_year",
        "session",
        "assessor",
        "teacher",
        "comments",
        "date_recorded"
    ])
    
    # Write data
    for assessment in assessments:
        teacher_name = assessment.assigned_teacher.username if assessment.assigned_teacher else "N/A"
        writer.writerow([
            assessment.student.student_number,
            assessment.student.full_name(),
            assessment.category,
            assessment.subject,
            assessment.score,
            assessment.max_score,
            f"{assessment.get_percentage():.2f}",
            assessment.term,
            assessment.academic_year,
            assessment.session,
            assessment.assessor,
            teacher_name,
            assessment.comments,
            assessment.date_recorded.strftime("%Y-%m-%d %H:%M:%S")
        ])
    
    # Convert to bytes
    mem = io.BytesIO()
    mem.write(si.getvalue().encode("utf-8"))
    mem.seek(0)
    
    return send_file(
        mem,
        as_attachment=True,
        download_name="assessments_export.csv",
        mimetype="text/csv"
    )
    
@app.route("/export/excel/assessment-template/<int:student_id>")
@login_required
def export_assessment_template(student_id):
    """Export student data to the assessment template Excel format"""
    student = Student.query.get_or_404(student_id)
    
    # Get all assessments for this student
    assessments = Assessment.query.filter_by(student_id=student.id, archived=False).all()
    
    # Create a template path
    template_path = os.path.join(app.config['TEMPLATE_FOLDER'], 'student_template.xlsx')
    
    # If template doesn't exist, create a default one
    if not os.path.exists(template_path):
        # You'll need to copy the actual template file here
        # For now, we'll create a placeholder
        flash("Template file not found. Please upload the template first.", "warning")
        return redirect(url_for('student_view', student_id=student_id))
    
    # Create output filename
    output_filename = f"{student.student_number}_{student.last_name}_assessment.xlsx"
    output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
    
    try:
        # Initialize template updater
        updater = AssessmentTemplateUpdater(template_path)
        updater.load_template()
        
        # Get student data in template format
        subject = None
        if current_user.is_teacher() and current_user.subject:
            subject = current_user.subject
        student_data = student.to_template_dict(subject)
        
        # Add student to template
        updater.add_student(10, student_data)
        
        # Save the updated workbook
        updater.save_workbook(output_path)
        
        # Send file to user
        return send_file(
            output_path,
            as_attachment=True,
            download_name=output_filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except PermissionError:
        flash("The Excel template file is currently open in another program (like Excel). Please close it and try the export again.", "danger")
        return redirect(url_for('student_view', student_id=student_id))
    except Exception as e:
        app.logger.error(f"Error exporting assessment template: {str(e)}")
        flash(f"Error exporting assessment template: {str(e)}", "danger")
        return redirect(url_for('student_view', student_id=student_id))

# Add a route to upload template
@app.route("/upload/template", methods=["GET", "POST"])
@login_required
@admin_required
def upload_template():
    """Upload assessment template Excel file"""
    if request.method == 'POST':
        if 'template_file' not in request.files:
            flash('No file selected', 'danger')
            return redirect(request.url)
        
        file = request.files['template_file']
        if file.filename == '':
            flash('No file selected', 'danger')
            return redirect(request.url)
        
        if file and file.filename.endswith('.xlsx'):
            filename = secure_filename('student_template.xlsx')
            filepath = os.path.join(app.config['TEMPLATE_FOLDER'], filename)
            file.save(filepath)
            flash('Template uploaded successfully', 'success')
            return redirect(url_for('dashboard'))
    
    return render_template("upload_template.html")

@app.route("/export/student/<int:student_id>/csv")
@login_required
def export_student_csv(student_id):
    student = Student.query.get_or_404(student_id)
    
    subject = request.args.get('subject')
    
    # Filter assessments by subject if specified
    assessments = student.assessments
    if subject:
        assessments = [a for a in assessments if a.subject == subject]
    
    # Create CSV in memory
    si = io.StringIO()
    writer = csv.writer(si)
    
    # Write header
    writer.writerow([
        "category",
        "subject",
        "class",
        "score",
        "max_score",
        "percentage",
        "grade",
        "term",
        "academic_year",
        "session",
        "assessor",
        "teacher",
        "comments",
        "date_recorded"
    ])
    
    # Write data
    for assessment in assessments:
        teacher_name = assessment.assigned_teacher.username if assessment.assigned_teacher else "N/A"
        writer.writerow([
            assessment.category,
            assessment.subject,
            assessment.class_name,
            assessment.score,
            assessment.max_score,
            f"{assessment.get_percentage():.2f}",
            assessment.get_grade_letter(),
            assessment.term,
            assessment.academic_year,
            assessment.session,
            assessment.assessor,
            teacher_name,
            assessment.comments,
            assessment.date_recorded.strftime("%Y-%m-%d %H:%M:%S")
        ])
    
    # Convert to bytes
    mem = io.BytesIO()
    mem.write(si.getvalue().encode("utf-8"))
    mem.seek(0)
    
    subject_str = f"_{subject}" if subject else ""
    filename = f"{student.student_number}_{student.last_name}_assessments{subject_str}.csv"
    
    return send_file(
        mem,
        as_attachment=True,
        download_name=filename,
        mimetype="text/csv"
    )

# -------------------------
# Excel Export/Import Routes
# -------------------------
@app.route("/export/excel/student/<int:student_id>")
@login_required
def export_student_excel(student_id):
    """Export single student to Excel template"""
    student = Student.query.get_or_404(student_id)
    
    subject = request.args.get('subject')
    
    # Get or create template path
    template_path = os.path.join(app.config['TEMPLATE_FOLDER'], 'student_template.xlsx')
    
    # Create default template if it doesn't exist
    if not os.path.exists(template_path):
        create_default_template(template_path)
        flash("Default template created. You can customize it in templates_excel folder.", "info")
    
    # Create output file
    subject_str = f"_{subject}" if subject else ""
    output_filename = f"{student.student_number}_{student.last_name}_report{subject_str}.xlsx"
    output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
    
    try:
        # Get settings
        settings = Setting.query.first()
        
        # Initialize template updater
        updater = AssessmentTemplateUpdater(template_path)
        updater.load_template()
        
        # Update school info
        if settings:
            updater.update_school_info(
                subject=subject or student.study_area,
                term_year=f"{settings.current_term} {settings.current_academic_year}",
                form=student.class_name
            )
        
        # Get student data in template format
        student_data = student.to_template_dict(subject)
        
        # Add student to template
        updater.add_student(10, student_data)
        
        # Save the updated workbook
        updater.save_workbook(output_path)
        
        # Send file to user
        return send_file(
            output_path,
            as_attachment=True,
            download_name=output_filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except PermissionError:
        flash("The Excel template file is currently open in another program (like Excel). Please close it and try the export again.", "danger")
        return redirect(url_for('student_view', student_id=student_id))
    except Exception as e:
        flash(f"Error exporting to Excel: {str(e)}", "danger")
        return redirect(url_for('student_view', student_id=student_id))

@app.route("/export/excel/all-students")
@login_required
def export_all_students_excel():
    """Export all students to Excel template"""
    subject = request.args.get('subject')
    class_name = request.args.get('class')
    
    # Filter students based on subject and class
    query = Student.query
    if subject:
        # Get students who have assessments in this subject
        subquery = db.session.query(Assessment.student_id).filter(Assessment.subject == subject).distinct()
        query = query.filter(Student.id.in_(subquery))
    if class_name:
        query = query.filter_by(class_name=class_name)
    
    students = query.order_by(Student.last_name, Student.first_name).all()
    
    # Get or create template path
    template_path = os.path.join(app.config['TEMPLATE_FOLDER'], 'student_template.xlsx')
    
    # Create default template if it doesn't exist
    if not os.path.exists(template_path):
        create_default_template(template_path)
    
    # Create output file
    subject_str = subject or "all_subjects"
    class_str = class_name or "all_classes"
    output_filename = f"students_{subject_str}_{class_str}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
    
    try:
        # Get settings
        settings = Setting.query.first()
        
        # Initialize template updater
        updater = AssessmentTemplateUpdater(template_path)
        updater.load_template()
        
        # Update school info
        form = class_name or "All Classes"
        subj = subject or "All Subjects"
        
        if settings:
            updater.update_school_info(
                subject=subj,
                term_year=f"{settings.current_term} {settings.current_academic_year}",
                form=form
            )
        
        # Get all students data in template format
        students_data = [student.to_template_dict() for student in students]
        
        # Add all students to template
        updater.add_students_batch(students_data)
        
        # Save the updated workbook
        updater.save_workbook(output_path)
        
        # Send file to user
        return send_file(
            output_path,
            as_attachment=True,
            download_name=output_filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except PermissionError:
        flash("The Excel template file is currently open in another program (like Excel). Please close it and try the export again.", "danger")
        return redirect(url_for('students'))
    except Exception as e:
        flash(f"Error exporting to Excel: {str(e)}", "danger")
        return redirect(url_for('students'))

@app.route("/export/assessments/excel")
@login_required
def export_assessments_excel():
    """Export filtered assessments to Excel"""
    from openpyxl import Workbook
    
    subject = request.args.get('subject', '')
    class_name = request.args.get('class', '')
    category = request.args.get('category', '')
    
    # Build query based on filters
    query = Assessment.query.filter_by(archived=False)
    if subject:
        query = query.filter_by(subject=subject)
    if class_name:
        query = query.filter_by(class_name=class_name)
    if category:
        query = query.filter_by(category=category)
    
    assessments = query.order_by(Assessment.date_recorded.desc()).all()
    
    # Create output file
    filters = []
    if subject: filters.append(subject)
    if class_name: filters.append(class_name)
    if category: filters.append(category)
    filter_str = "_".join(filters) if filters else "all"
    output_filename = f"assessments_{filter_str}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    output_path = os.path.join(app.config['UPLOAD_FOLDER'], output_filename)
    
    try:
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Assessments"
        
        # Headers
        headers = [
            "Student Number", "Student Name", "Subject", "Category", 
            "Score", "Max Score", "Percentage", "Grade", "Class", 
            "Term", "Academic Year", "Session", "Assessor", "Date Recorded"
        ]
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Data
        for row, assessment in enumerate(assessments, 2):
            teacher_name = assessment.assigned_teacher.username if assessment.assigned_teacher else "N/A"
            ws.cell(row=row, column=1, value=assessment.student.student_number)
            ws.cell(row=row, column=2, value=assessment.student.full_name())
            ws.cell(row=row, column=3, value=assessment.subject)
            ws.cell(row=row, column=4, value=assessment.category)
            ws.cell(row=row, column=5, value=assessment.score)
            ws.cell(row=row, column=6, value=assessment.max_score)
            ws.cell(row=row, column=7, value=round(assessment.get_percentage(), 2))
            ws.cell(row=row, column=8, value=assessment.get_grade_letter())
            ws.cell(row=row, column=9, value=assessment.class_name)
            ws.cell(row=row, column=10, value=assessment.term)
            ws.cell(row=row, column=11, value=assessment.academic_year)
            ws.cell(row=row, column=12, value=assessment.session)
            ws.cell(row=row, column=13, value=assessment.assessor)
            ws.cell(row=row, column=14, value=assessment.date_recorded.strftime("%Y-%m-%d %H:%M:%S"))
        
        # Save
        wb.save(output_path)
        
        # Send file
        return send_file(
            output_path,
            as_attachment=True,
            download_name=output_filename,
            mimetype="application/vnd/openxmlformats-officedocument.spreadsheetml.sheet"
        )
    except Exception as e:
        flash(f"Error exporting to Excel: {str(e)}", "danger")
        return redirect(url_for('assessments_list'))

@app.route("/import/excel", methods=["GET", "POST"])
@login_required
def import_excel():
    """Bulk import assessments from Excel file"""
    form = BulkImportForm()
    
    if form.validate_on_submit():
        file = form.excel_file.data
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        # Save uploaded file
        file.save(filepath)
        
        try:
            # Import assessments
            importer = ExcelBulkImporter(filepath)
            assessments_data = importer.import_assessments()
            
            # Process and save assessments
            success_count = 0
            error_count = 0
            errors = []
            
            for data in assessments_data:
                try:
                    # Find student
                    student = Student.query.filter_by(
                        student_number=data['student_number']
                    ).first()
                    
                    if not student:
                        errors.append(f"Student {data['student_number']} not found")
                        error_count += 1
                        continue
                    
                    # Check if assessment already exists for this student, category, subject, term, academic_year, session
                    existing_assessment = Assessment.query.filter_by(
                        student_id=student.id,
                        category=data['category'],
                        subject=data['subject'],
                        term=data['term'],
                        academic_year=data.get('academic_year'),
                        session=data['session']
                    ).first()
                    
                    if existing_assessment:
                        errors.append(f"Assessment for {data['category']} in {data['subject']} already exists for student {data['student_number']} in the same term, academic year, and session")
                        error_count += 1
                        continue
                    
                    # Create assessment
                    assessment = Assessment(
                        student=student,
                        category=data['category'],
                        subject=data['subject'],
                        score=float(data['score']),
                        max_score=float(data['max_score']),
                        term=data['term'],
                        academic_year=data.get('academic_year'),
                        session=data['session'],
                        assessor=data['assessor'],
                        teacher_id=current_user.id if hasattr(current_user, 'is_teacher') and current_user.is_teacher() else None,
                        comments=data['comments']
                    )
                    db.session.add(assessment)
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"Row error: {str(e)}")
                    error_count += 1
            
            # Commit all changes
            db.session.commit()
            
            # Clean up uploaded file
            os.remove(filepath)
            
            # Show results
            flash(f"Successfully imported {success_count} assessments", "success")
            if error_count > 0:
                flash(f"{error_count} errors occurred: {'; '.join(errors[:5])}", "warning")
            
            return redirect(url_for('assessments_list'))
            
        except Exception as e:
            db.session.rollback()
            if os.path.exists(filepath):
                os.remove(filepath)
            flash(f"Error importing file: {str(e)}", "danger")
    
    return render_template("import_excel.html", form=form)

@app.route("/download/template/<template_type>")
@login_required
def download_template(template_type):
    """Download Excel template"""
    if template_type == "student":
        template_path = os.path.join(app.config['TEMPLATE_FOLDER'], 'student_template.xlsx')
        filename = "student_assessment_template.xlsx"
    elif template_type == "import":
        template_path = os.path.join(app.config['TEMPLATE_FOLDER'], 'import_template.xlsx')
        filename = "bulk_import_template.xlsx"
    elif template_type == "student_import":
        template_path = os.path.join(app.config['TEMPLATE_FOLDER'], 'student_import_template.xlsx')
        filename = "student_bulk_import_template.xlsx"
    else:
        abort(404)
    
    # For import template, use the existing one, do not create default
    if template_type == "import" and not os.path.exists(template_path):
        flash("Import template not found. Please contact administrator.", "danger")
        return redirect(url_for('import_excel'))
    
    # Create template if it doesn't exist (for student template)
    if template_type == "student" and not os.path.exists(template_path):
        create_default_template(template_path)
    elif template_type == "student_import" and not os.path.exists(template_path):
        create_student_import_template(template_path)
    
    return send_file(
        template_path,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

@app.route("/download/question_template")
@login_required
@teacher_required
def download_question_template():
    """Download question import template"""
    template_path = os.path.join(app.config['TEMPLATE_FOLDER'], 'question_import_template.xlsx')
    
    # Create template if it doesn't exist
    if not os.path.exists(template_path):
        create_question_import_template(template_path)
    
    return send_file(
        template_path,
        as_attachment=True,
        download_name="question_bulk_import_template.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# -------------------------
# Error Handlers
# -------------------------
@app.errorhandler(403)
def forbidden(e):
    return render_template("403.html"), 403

@app.errorhandler(404)
def not_found(e):
    return render_template("404.html"), 404

@app.errorhandler(500)
def internal_error(e):
    db.session.rollback()
    return render_template("500.html"), 500

# -------------------------
# Context Processors
# -------------------------
@app.context_processor
def inject_config():
    """Make config values available in templates"""
    return {
        'CATEGORY_LABELS': app.config['CATEGORY_LABELS'],
        'ASSESSMENT_WEIGHTS': app.config['ASSESSMENT_WEIGHTS'],
        'LEARNING_AREAS': app.config['LEARNING_AREAS'],
        'CLASS_LEVELS': app.config['CLASS_LEVELS']
    }

# -------------------------
# Run Application
# -------------------------
if __name__ == "__main__":
    print("\n" + "="*60)
    print("Student Assessment Management System")
    print("="*60)
    print(f"Environment: {env}")
    print(f"Database: {app.config['SQLALCHEMY_DATABASE_URI']}")
    print(f"Access at: http://127.0.0.1:5000")
    print("="*60 + "\n")
    
    app.run(
        debug=app.config.get('DEBUG', True), 
        host='127.0.0.1', 
        port=5000,
        use_reloader=True
    )