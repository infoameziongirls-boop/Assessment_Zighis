# template_updater.py
import os
import tempfile
import shutil
from openpyxl import load_workbook

class AssessmentTemplateUpdater:
    def __init__(self, template_path):
        """
        Initialize with the path to the template Excel file
        
        Args:
            template_path (str): Path to the student_template.xlsx file
        """
        self.template_path = template_path
        self.wb = None
        self.ws = None
        self.start_row = 10  # Student data starts at row 10
        self.max_students = 111  # From row 10 to row 120
        
        # Define column indices based on the template
        self.columns = {
            'serial': 'A',
            'ref_id': 'A',  # Reference Number in A10
            'name': 'B',  # Surname Firstname Othername in B10
            'student_number': 'C',  # Student Number in C10
            'study_area': 'D',  # Study Area in D10
            'ica1': 'E',  # Individual Class Assessment 1
            'ica2': 'F',  # Individual Class Assessment 2
            'ica_total': 'G',  # SUB TOTAL (I.C.A.)
            'icp1': 'H',  # Individual Class Project 1
            'icp2': 'I',  # Individual Class Project 2
            'icp_total': 'J',  # SUB TOTAL TEST(C.P)
            'gp1': 'K',  # Group Project/Research 1
            'gp2': 'L',  # Group Project/Research 2
            'gp_total': 'M',  # SUB TOTAL (G.P)
            'practical': 'N',  # PRACTICAL PORTFOLIO
            'mid_term': 'O',  # MID TERM EXAMS
            'total_class': 'P',  # TOTAL CLASS SCORE
            'percent_class': 'Q',  # 100% OF TOTAL CLASS SCORE
            'avg_class': 'R',  # AVG. CLASS SC.
            'end_term': 'S',  # END OF TERM EXAMS
            'avg_exam': 'T',  # AVG. EXAMS SC.
            'total_50_50': 'U',  # TOTAL 50 + 50
            'weighted_result': 'V',  # Weighted Result (percentage)
            'gpa': 'W',  # GPA
            'grade': 'X'  # Grade
        }
        
    def load_template(self):
        """Load the Excel template"""
        # Copy template to temp file to avoid lock issues
        temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(temp_fd)  # Close the file descriptor
        shutil.copy2(self.template_path, temp_path)
        self.temp_path = temp_path
        self.wb = load_workbook(temp_path)
        self.ws = self.wb['ASSESSMENT TEMPLATE']
        return self
    
    def save_workbook(self, output_path):
        """Save the updated workbook"""
        self.wb.save(output_path)
        
    def add_student(self, row, student_data):
        """
        Add or update a student's data at the specified row
        
        Args:
            row (int): Row number (starting from 10)
            student_data (dict): Dictionary containing student data
        """
        if row < self.start_row or row > self.start_row + self.max_students:
            raise ValueError(f"Row must be between {self.start_row} and {self.start_row + self.max_students}")
        
        # Basic student info
        if 'name' in student_data:
            self.ws[f"{self.columns['name']}{row}"] = student_data['name']
        if 'ref_id' in student_data:
            self.ws[f"{self.columns['ref_id']}{row}"] = student_data['ref_id']
        if 'student_number' in student_data:
            self.ws[f"{self.columns['student_number']}{row}"] = student_data['student_number']
        if 'study_area' in student_data:
            self.ws[f"{self.columns['study_area']}{row}"] = student_data['study_area']
        
        # Individual Class Assessment marks
        if 'ica1' in student_data:
            self.ws[f"{self.columns['ica1']}{row}"] = self._parse_mark(student_data['ica1'])
        if 'ica2' in student_data:
            self.ws[f"{self.columns['ica2']}{row}"] = self._parse_mark(student_data['ica2'])
            
        # Individual Class Project marks
        if 'icp1' in student_data:
            self.ws[f"{self.columns['icp1']}{row}"] = self._parse_mark(student_data['icp1'])
        if 'icp2' in student_data:
            self.ws[f"{self.columns['icp2']}{row}"] = self._parse_mark(student_data['icp2'])
            
        # Group Project marks
        if 'gp1' in student_data:
            self.ws[f"{self.columns['gp1']}{row}"] = self._parse_mark(student_data['gp1'])
        if 'gp2' in student_data:
            self.ws[f"{self.columns['gp2']}{row}"] = self._parse_mark(student_data['gp2'])
            
        # Other assessments
        if 'practical' in student_data:
            self.ws[f"{self.columns['practical']}{row}"] = self._parse_mark(student_data['practical'])
        if 'mid_term' in student_data:
            self.ws[f"{self.columns['mid_term']}{row}"] = self._parse_mark(student_data['mid_term'])
        if 'end_term' in student_data:
            self.ws[f"{self.columns['end_term']}{row}"] = self._parse_mark(student_data['end_term'])
        
        # Update formulas if they were overwritten
        self._restore_formulas(row)
        
    def add_students_batch(self, students_data):
        """
        Add multiple students starting from the first available row
        
        Args:
            students_data (list): List of student data dictionaries
        """
        current_row = self.start_row
        
        for i, student_data in enumerate(students_data):
            if current_row > self.start_row + self.max_students:
                print(f"Warning: Maximum student capacity ({self.max_students}) reached.")
                break
            
            # Add serial number
            self.ws[f"{self.columns['serial']}{current_row}"] = i + 1
            
            # Add student data
            self.add_student(current_row, student_data)
            current_row += 1
            
        # Update total students count formula (cell C3)
        self._update_total_students()
        
    def _parse_mark(self, mark):
        """
        Parse mark value, handling special cases like 'E' for excused
        
        Args:
            mark: Can be int, float, str, or None
            
        Returns:
            Appropriate value for Excel cell
        """
        if mark is None:
            return None
            
        if isinstance(mark, str):
            if mark.upper() == 'E':
                return 'E'  # Excused
            try:
                # Try to convert string numbers to float
                return float(mark)
            except ValueError:
                return mark
        return mark
    
    def _restore_formulas(self, row):
        """Restore formulas for calculated columns if they were overwritten"""
        # G column: =MIN(100, (SUM(E{row}:F{row})))
        self.ws[f"{self.columns['ica_total']}{row}"].value = f'=MIN(100, (SUM(E{row}:F{row})))'
        
        # J column: =MIN(100,(SUM(H{row}:I{row})))
        self.ws[f"{self.columns['icp_total']}{row}"].value = f'=MIN(100,(SUM(H{row}:I{row})))'
        
        # M column: =MIN(100,(SUM(K{row}:L{row})))
        self.ws[f"{self.columns['gp_total']}{row}"].value = f'=MIN(100,(SUM(K{row}:L{row})))'
        
        # P column: =MIN(500, (SUM(G{row},J{row},M{row},N{row},O{row})))
        self.ws[f"{self.columns['total_class']}{row}"].value = f'=MIN(500, (SUM(G{row},J{row},M{row},N{row},O{row})))'
        
        # Q column: =P{row}/500*100
        self.ws[f"{self.columns['percent_class']}{row}"].value = f'=P{row}/500*100'
        
        # R column: =MIN(50, (ROUNDUP(SUM(Q{row})/2,0)))
        self.ws[f"{self.columns['avg_class']}{row}"].value = f'=MIN(50, (ROUNDUP(SUM(Q{row})/2,0)))'
        
        # T column: =MIN(50, (ROUNDUP(SUM(S{row})/2,0)))
        self.ws[f"{self.columns['avg_exam']}{row}"].value = f'=MIN(50, (ROUNDUP(SUM(S{row})/2,0)))'
        
        # U column: =MIN(100, (SUM(R{row},T{row})))
        self.ws[f"{self.columns['total_50_50']}{row}"].value = f'=MIN(100, (SUM(R{row},T{row})))'
        
        # V column: Weighted result (percentage) - pick value from U and add percent
        self.ws[f"{self.columns['weighted_result']}{row}"].value = f'=U{row}'
        
        # W column: GPA formula
        self.ws[f"{self.columns['gpa']}{row}"].value = (
            f'=IF(U{row}>=80,"4.0",IF(U{row}>=70,"3.5",IF(U{row}>=65,"3.0",'
            f'IF(U{row}>=60,"2.5",IF(U{row}>=55,"2.0",IF(U{row}>=50,"1.5",'
            f'IF(U{row}>=45,"1.0",IF(U{row}>=40,"0.5",IF(U{row}<40,"0.0")))))))))'
        )
        
        # X column: Grade formula
        self.ws[f"{self.columns['grade']}{row}"].value = (
            f'=IF(U{row}>=80,"A1",IF(U{row}>=70,"B2",IF(U{row}>=65,"B3",'
            f'IF(U{row}>=60,"C4",IF(U{row}>=55,"C5",IF(U{row}>=50,"C6",'
            f'IF(U{row}>=45,"D7",IF(U{row}>=40,"E8",IF(U{row}<40,"F9")))))))))'
        )
    
    def _update_total_students(self):
        """Update the formula for total students count"""
        # Cell C3 has =COUNTA(B10:B110)
        self.ws['C3'].value = '=COUNTA(B10:B110)'
    
    def clear_student_data(self, start_row=None, end_row=None):
        """
        Clear student data from specified rows
        
        Args:
            start_row (int): Starting row (default: 10)
            end_row (int): Ending row (default: 120)
        """
        if start_row is None:
            start_row = self.start_row
        if end_row is None:
            end_row = self.start_row + self.max_students
            
        for row in range(start_row, end_row + 1):
            # Clear student info columns
            for col in ['B', 'C', 'D']:
                self.ws[f"{col}{row}"].value = None
            
            # Clear assessment columns but keep formulas
            input_columns = ['E', 'F', 'H', 'I', 'K', 'L', 'N', 'O', 'S']
            for col in input_columns:
                self.ws[f"{col}{row}"].value = 0
            
            # Restore formulas
            self._restore_formulas(row)
    
    def get_student_data(self, row):
        """
        Get student data from a specific row
        
        Args:
            row (int): Row number
            
        Returns:
            dict: Student data
        """
        data = {}
        for key, col in self.columns.items():
            cell_value = self.ws[f"{col}{row}"].value
            if cell_value is not None:
                data[key] = cell_value
        return data
    
    def get_all_students(self):
        """
        Get data for all students with entries
        
        Returns:
            list: List of student data dictionaries
        """
        students = []
        for row in range(self.start_row, self.start_row + self.max_students + 1):
            name = self.ws[f"{self.columns['name']}{row}"].value
            if name:  # Only include rows with names
                students.append(self.get_student_data(row))
        return students
    
    def update_school_info(self, subject=None, term_year=None, form=None):
        """Update school information at the top of the sheet"""
        if subject:
            self.ws['B2'] = subject
        if term_year:
            self.ws['B3'] = term_year
        if form:
            self.ws['B4'] = form