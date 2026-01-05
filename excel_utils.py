"""
Excel Template Handler
Handles reading from and writing to Excel templates while preserving formatting
"""
import os
import tempfile
import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Alignment
from datetime import datetime


class ExcelTemplateHandler:
    """Handle Excel template operations"""
    
    def __init__(self, template_path):
        """
        Initialize with path to Excel template
        
        Args:
            template_path: Path to the Excel template file
        """
        self.template_path = template_path
        
    def load_template(self):
        """Load the Excel template workbook"""
        if not os.path.exists(self.template_path):
            raise FileNotFoundError(f"Template not found: {self.template_path}")
        
        # Copy template to temp file to avoid lock issues
        temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
        os.close(temp_fd)
        shutil.copy2(self.template_path, temp_path)
        self.temp_path = temp_path
        return load_workbook(temp_path)
    
    def export_student_to_template(self, student, assessments, output_path, config):
        """
        Export student data to Excel template
        
        Args:
            student: Student object
            assessments: List of Assessment objects
            output_path: Where to save the filled template
            config: App config with category labels and weights
        """
        # Load template
        wb = self.load_template()
        ws = wb.active  # Use the first sheet
        
        # Fill in student information (adjust cell references based on your template)
        self._write_student_info(ws, student)
        
        # Fill in assessments by category
        self._write_assessments(ws, assessments, config)
        
        # Calculate and write summary statistics
        self._write_summary(ws, assessments, config)
        
        # Save the filled template
        wb.save(output_path)
        return output_path
    
    def export_all_students_to_template(self, students, output_path, config):
        """
        Export all students to a summary Excel template
        
        Args:
            students: List of Student objects
            output_path: Where to save the filled template
            config: App config
        """
        wb = self.load_template()
        ws = wb.active
        
        # Starting row for data (adjust based on your template)
        start_row = 5  # Assumes rows 1-4 are headers
        
        for idx, student in enumerate(students):
            row = start_row + idx
            
            # Write student data (adjust column indices based on your template)
            ws.cell(row=row, column=1, value=student.student_number)
            ws.cell(row=row, column=2, value=student.full_name())
            ws.cell(row=row, column=3, value=student.learning_area if student.learning_area else "")
            
            # Calculate summary for each category
            summary = student.get_assessment_summary()
            
            # Write category averages (adjust columns as needed)
            col = 4
            for cat in ["IA", "IPA", "PP", "MSE", "ETE"]:
                if cat in summary:
                    ws.cell(row=row, column=col, value=summary[cat]["avg_percent"])
                else:
                    ws.cell(row=row, column=col, value="")
                col += 1
            
            # Write final grade
            final = student.calculate_final_grade()
            ws.cell(row=row, column=col, value=final if final else "")
        
        wb.save(output_path)
        return output_path
    
    def _write_student_info(self, ws, student):
        """Write student information to specific cells"""
        # Adjust these cell references based on your template
        ws['B2'] = student.student_number
        ws['B3'] = student.full_name()
        ws['B4'] = student.learning_area if student.learning_area else ""
        ws['B5'] = datetime.now().strftime('%Y-%m-%d')
    
    def _write_assessments(self, ws, assessments, config):
        """Write assessment data to the template"""
        # Group assessments by category
        by_category = {}
        for assessment in assessments:
            if assessment.category not in by_category:
                by_category[assessment.category] = []
            by_category[assessment.category].append(assessment)
        
        # Define starting rows for each category (adjust based on your template)
        category_rows = {
            "IA": 8,    # Individual Assessments start at row 8
            "IPA": 13,  # Individual Projects start at row 13
            "PP": 18,   # Practical Portfolio start at row 18
            "MSE": 23,  # Mid-Semester Exams start at row 23
            "ETE": 28   # End-of-Term Exams start at row 28
        }
        
        # Write each category's assessments
        for category, start_row in category_rows.items():
            if category in by_category:
                assessments_list = by_category[category]
                
                for idx, assessment in enumerate(assessments_list):
                    row = start_row + idx
                    
                    # Write assessment data (adjust columns based on template)
                    ws.cell(row=row, column=1, value=assessment.subject or "")
                    ws.cell(row=row, column=2, value=assessment.score)
                    ws.cell(row=row, column=3, value=assessment.max_score)
                    ws.cell(row=row, column=4, value=assessment.get_percentage())
                    ws.cell(row=row, column=5, value=assessment.term or "")
                    ws.cell(row=row, column=6, value=assessment.session or "")
                    ws.cell(row=row, column=7, value=assessment.assessor or "")
                    ws.cell(row=row, column=8, value=assessment.comments or "")
    
    def _write_summary(self, ws, assessments, config):
        """Write summary statistics"""
        from collections import defaultdict
        
        # Calculate summary by category
        summary = defaultdict(lambda: {"total_score": 0, "total_max": 0, "count": 0})
        
        for assessment in assessments:
            cat = assessment.category
            summary[cat]["total_score"] += assessment.score
            summary[cat]["total_max"] += assessment.max_score
            summary[cat]["count"] += 1
        
        # Write summary to specific cells (adjust based on your template)
        summary_row = 35  # Starting row for summary section
        
        for idx, (cat, label) in enumerate(config['CATEGORY_LABELS'].items()):
            row = summary_row + idx
            
            if cat in summary and summary[cat]["total_max"] > 0:
                avg = (summary[cat]["total_score"] / summary[cat]["total_max"]) * 100
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=summary[cat]["count"])
                ws.cell(row=row, column=3, value=f"{avg:.2f}%")
            else:
                ws.cell(row=row, column=1, value=label)
                ws.cell(row=row, column=2, value=0)
                ws.cell(row=row, column=3, value="N/A")
        
        # Calculate and write final weighted grade
        final = 0.0
        weight_used = 0.0
        
        for cat, weight in config['ASSESSMENT_WEIGHTS'].items():
            if cat in summary and summary[cat]["total_max"] > 0:
                avg = (summary[cat]["total_score"] / summary[cat]["total_max"]) * 100
                final += avg * weight
                weight_used += weight
        
        if weight_used > 0:
            ws.cell(row=summary_row + 6, column=1, value="Final Weighted Grade:")
            ws.cell(row=summary_row + 6, column=3, value=f"{final:.2f}%")


class ExcelBulkImporter:
    """Handle bulk import of assessments from Excel"""
    
    def __init__(self, file_path):
        """
        Initialize with path to Excel file
        
        Args:
            file_path: Path to the Excel file to import
        """
        self.file_path = file_path
    
    def import_assessments(self, start_row=2):
        """
        Import assessments from Excel file
        
        Args:
            start_row: Row number where data starts (default 2, assumes row 1 is header)
            
        Returns:
            List of dictionaries containing assessment data
        """
        wb = load_workbook(self.file_path, data_only=True)
        ws = wb.active
        
        assessments = []
        
        # Read data starting from start_row
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            # Skip empty rows
            if not any(row):
                continue
            
            # Map columns to fields (adjust indices based on your template)
            assessment_data = {
                'student_number': row[0],      # Column A
                'category': row[1],            # Column B
                'subject': row[2],             # Column C
                'score': row[3],               # Column D
                'max_score': row[4],           # Column E
                'term': row[5],                # Column F
                'session': row[6],             # Column G
                'assessor': row[7],            # Column H
                'comments': row[8] if len(row) > 8 else ""  # Column I
            }
            
            # Validate required fields
            if assessment_data['student_number'] and assessment_data['score'] is not None:
                assessments.append(assessment_data)
        
        wb.close()
        return assessments


class StudentBulkImporter:
    """Handle bulk import of students from Excel"""
    
    def __init__(self, file_path):
        """
        Initialize with path to Excel file
        
        Args:
            file_path: Path to the Excel file to import
        """
        self.file_path = file_path
    
    def import_students(self, start_row=2):
        """
        Import students from Excel file
        
        Args:
            start_row: Row number where data starts (default 2, assumes row 1 is header)
            
        Returns:
            List of dictionaries containing student data
        """
        wb = load_workbook(self.file_path, data_only=True)
        ws = wb.active
        
        students = []
        
        # Read data starting from start_row
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            # Skip empty rows
            if not any(row):
                continue
            
            # Map columns to fields (adjust indices based on your template)
            student_data = {
                'student_number': row[0],      # Column A
                'first_name': row[1],          # Column B
                'last_name': row[2],           # Column C
                'middle_name': row[3] if len(row) > 3 else None,  # Column D
                'class_name': row[4] if len(row) > 4 else None,   # Column E
                'study_area': row[5] if len(row) > 5 else None    # Column F
            }
            
            # Validate required fields
            if student_data['student_number'] and student_data['first_name'] and student_data['last_name']:
                students.append(student_data)
        
        wb.close()
        return students


class QuestionBulkImporter:
    """Handle bulk import of questions from Excel"""
    
    def __init__(self, file_path):
        """
        Initialize with path to Excel file
        
        Args:
            file_path: Path to the Excel file to import
        """
        self.file_path = file_path
    
    def import_questions(self, start_row=2):
        """
        Import questions from Excel file
        
        Args:
            start_row: Row number where data starts (default 2, assumes row 1 is header)
            
        Returns:
            List of dictionaries containing question data
        """
        wb = load_workbook(self.file_path, data_only=True)
        ws = wb.active
        
        questions = []
        
        # Read data starting from start_row
        for row in ws.iter_rows(min_row=start_row, values_only=True):
            # Skip empty rows
            if not any(row):
                continue
            
            # Map columns to fields
            question_data = {
                'question_text': row[0],       # Column A
                'question_type': row[1],       # Column B: mcq, true_false, short_answer
                'option_a': row[2] if len(row) > 2 else None,      # Column C
                'option_b': row[3] if len(row) > 3 else None,      # Column D
                'option_c': row[4] if len(row) > 4 else None,      # Column E
                'option_d': row[5] if len(row) > 5 else None,      # Column F
                'correct_answer': row[6],     # Column G
                'difficulty': row[7] if len(row) > 7 else 'medium', # Column H
                'explanation': row[8] if len(row) > 8 else None    # Column I
            }
            
            # Validate required fields
            if question_data['question_text'] and question_data['question_type'] and question_data['correct_answer']:
                # Process options based on type
                if question_data['question_type'].lower() == 'mcq':
                    question_data['options'] = [question_data['option_a'], question_data['option_b'], 
                                              question_data['option_c'], question_data['option_d']]
                else:
                    question_data['options'] = None
                
                questions.append(question_data)
        
        wb.close()
        return questions


def create_default_template(output_path):
    """
    Create a default Excel template if none exists
    
    Args:
        output_path: Where to save the template
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Assessment"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Student Info Section
    ws['A1'] = "STUDENT INFORMATION"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['A1'].font = Font(bold=True, color="FFFFFF", size=14)
    
    ws['A2'] = "Student Number:"
    ws['A3'] = "Full Name:"
    ws['A4'] = "Learning Area:"
    ws['A5'] = "Report Date:"
    
    # Make labels bold
    for cell in ['A2', 'A3', 'A4', 'A5']:
        ws[cell].font = Font(bold=True)
    
    # Assessment Categories Headers
    categories = [
        ("A7", "INDIVIDUAL ASSESSMENTS (I.A)"),
        ("A12", "INDIVIDUAL PROJECTS (I.P.A)"),
        ("A17", "PRACTICAL PORTFOLIO (P.P)"),
        ("A22", "MID-SEMESTER EXAM (M.S.E)"),
        ("A27", "END-OF-TERM EXAM (E.T.E)")
    ]
    
    for cell, title in categories:
        ws[cell] = title
        ws[cell].font = Font(bold=True, size=12)
        ws[cell].fill = header_fill
        ws[cell].font = Font(bold=True, color="FFFFFF", size=12)
    
    # Column headers for each section
    col_headers = ["Subject", "Score", "Max Score", "Percentage", "Term", "Session", "Assessor", "Comments"]
    header_row = 8
    
    for _ in range(5):  # 5 categories
        for idx, header in enumerate(col_headers):
            cell = ws.cell(row=header_row, column=idx+1, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        header_row += 5
    
    # Summary Section
    ws['A34'] = "SUMMARY"
    ws['A34'].font = Font(bold=True, size=14)
    ws['A34'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    ws['A34'].font = Font(bold=True, color="FFFFFF", size=14)
    
    summary_headers = ["Category", "Count", "Average"]
    for idx, header in enumerate(summary_headers):
        cell = ws.cell(row=35, column=idx+1, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    
    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 30
    
    wb.save(output_path)
    return output_path


def create_student_import_template(output_path):
    """
    Create a student bulk import Excel template
    
    Args:
        output_path: Where to save the template
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Import"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Headers
    headers = ["Student Number", "First Name", "Last Name", "Middle Name", "Class", "Study Area"]
    
    for idx, header in enumerate(headers):
        cell = ws.cell(row=1, column=idx+1, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Sample data
    sample_data = [
        ["STU001", "John", "Doe", "Michael", "Grade 10", "Mathematics"],
        ["STU002", "Jane", "Smith", "", "Grade 9", "Science"],
        ["STU003", "Bob", "Johnson", "William", "Grade 11", "English"]
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data):
            ws.cell(row=row_idx, column=col_idx+1, value=value)
    
    # Set column widths
    column_widths = [15, 15, 15, 15, 10, 15]
    for idx, width in enumerate(column_widths):
        ws.column_dimensions[chr(65 + idx)].width = width
    
    wb.save(output_path)
    return output_path


def create_question_import_template(output_path):
    """
    Create Excel template for bulk question import
    
    Args:
        output_path: Where to save the template
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Question Import"
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Headers
    headers = ["Question Text", "Question Type", "Option A", "Option B", "Option C", "Option D", "Correct Answer", "Difficulty", "Explanation"]
    
    for idx, header in enumerate(headers):
        cell = ws.cell(row=1, column=idx+1, value=header)
        cell.font = header_font
        cell.fill = header_fill
    
    # Sample data
    sample_data = [
        ["What is the capital of France?", "mcq", "Paris", "London", "Berlin", "Madrid", "A", "easy", "Paris is the capital and largest city of France."],
        ["The Earth is round.", "true_false", "", "", "", "", "True", "easy", "Scientific evidence confirms the Earth is an oblate spheroid."],
        ["What is 2 + 2?", "short_answer", "", "", "", "", "4", "easy", "Basic arithmetic: 2 + 2 = 4."]
    ]
    
    for row_idx, row_data in enumerate(sample_data, start=2):
        for col_idx, value in enumerate(row_data):
            ws.cell(row=row_idx, column=col_idx+1, value=value)
    
    # Set column widths
    column_widths = [40, 15, 15, 15, 15, 15, 15, 10, 30]
    for idx, width in enumerate(column_widths):
        ws.column_dimensions[chr(65 + idx)].width = width
    
    wb.save(output_path)
    return output_path